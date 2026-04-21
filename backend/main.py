"""
API FastAPI para Vigilancia Epidemiológica — Brote Sarampión 2026
IGSS — Departamento de Medicina Preventiva — Sección de Epidemiología

Endpoints:
  POST /api/registro              → Guardar nuevo registro (público, desde formulario)
  GET  /api/registro/{id}         → Obtener registro individual (requiere API key)
  PUT  /api/registro/{id}         → Editar registro existente (requiere API key)
  GET  /api/registro/{id}/audit   → Historial de cambios (requiere API key)
  GET  /api/registros             → Listar registros (requiere API key)
  GET  /api/registros/count       → Contar registros
  POST /api/registros/upload      → Carga masiva desde Excel (requiere API key)
  GET  /api/export/excel          → Descargar como Excel (requiere API key)
  GET  /api/export/csv            → Descargar como CSV (requiere API key)
  GET  /api/health                → Health check
"""
import io
import csv
import re
import time
import hmac
import logging
from datetime import datetime

from fastapi import FastAPI, Request, HTTPException, Query, Header, UploadFile, File
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse, JSONResponse, FileResponse

from config import ALLOWED_ORIGINS, API_SECRET_KEY, PORT, RATE_LIMIT_SECONDS, MAX_UPLOAD_SIZE_MB
from database import (
    init_db, insert_registro, get_registros, get_count, check_duplicate,
    find_duplicate_21_dias,
    get_registro_by_id, update_registro, delete_registro, bulk_insert_registros,
    init_audit_table, log_changes, get_audit_trail, search_registros,
    COLUMNS, EDITABLE_COLUMNS,
)
from mspas_queue import (
    init_mspas_tables, save_credentials, get_credentials,
    enqueue_record, enqueue_all_pending, get_queue, get_queue_counts,
    approve_records, update_estado, get_approved_for_submission,
    mark_sent, mark_error, mark_duplicate, mark_possible_duplicate,
    try_claim_for_submission, recover_stuck_submissions, get_status_by_id,
)
from godata_queue import (
    init_godata_tables,
    save_godata_config, get_godata_config, get_godata_credentials,
    enqueue_pending_records as godata_enqueue_pending,
    get_queue as godata_get_queue,
    approve_records as godata_approve_records,
    try_claim_for_sync, mark_synced, mark_error as godata_mark_error,
    mark_duplicate as godata_mark_duplicate,
    get_sync_status as godata_get_sync_status,
    get_approved_for_sync, recover_stuck_syncs,
    get_next_visual_id, save_visual_id,
    mark_fase1_sent, mark_complete, mark_error_fase1, mark_error_fase2,
    get_fase1_pending, try_claim_for_fase2,
    requeue_for_update, unapprove_records,
)
from godata_client import GoDataClient
from godata_field_map import (
    map_record_to_godata, map_lab_samples_to_godata, validate_godata_payload,
    map_record_to_godata_fase1, map_record_to_godata_fase2,
)

logger = logging.getLogger(__name__)

# ─── App ──────────────────────────────────────────────────
app = FastAPI(
    title="Vigilancia Sarampión - IGSS API",
    description="API de registro de casos sospechosos de sarampión",
    version="2.0.0",
    docs_url=None,   # Disable Swagger in production
    redoc_url=None,
)

app.add_middleware(
    CORSMiddleware,
    allow_origins=ALLOWED_ORIGINS,
    allow_credentials=False,
    allow_methods=["GET", "POST", "PUT", "DELETE", "OPTIONS"],
    allow_headers=["*"],
)

# Rate limiting por IP (en memoria — single worker)
_rate_limiter: dict[str, float] = {}
_RATE_LIMIT_CLEANUP_INTERVAL = 300  # Limpiar cada 5 min
_last_cleanup = time.time()


def _cleanup_rate_limiter():
    """Elimina entradas viejas del rate limiter para evitar memory leak."""
    global _last_cleanup
    now = time.time()
    if now - _last_cleanup < _RATE_LIMIT_CLEANUP_INTERVAL:
        return
    _last_cleanup = now
    cutoff = now - 60
    keys_to_remove = [k for k, v in _rate_limiter.items() if v < cutoff]
    for k in keys_to_remove:
        del _rate_limiter[k]


# ─── Validación de datos ──────────────────────────────────
VALID_CLASIFICACIONES = {
    "SOSPECHOSO", "CONFIRMADO", "CONFIRMADO SARAMPIÓN", "CONFIRMADO SARAMPION",
    "CONFIRMADO RUBÉOLA", "CONFIRMADO RUBEOLA", "DESCARTADO", "PENDIENTE",
    "NO CUMPLE DEFINICIÓN", "NO CUMPLE DEFINICION",
    "NO CUMPLE DEFINICIÓN DE CASO", "NO CUMPLE DEFINICION DE CASO",
    "SUSPENDIDO", "CLÍNICO", "CLINICO", "FALSO",
    "ERROR DIAGNÓSTICO", "ERROR DIAGNOSTICO",
}
VALID_SEXO = {"M", "F", ""}
VALID_SI_NO = {"SI", "NO", "N/A", "N/S", "DESCONOCIDO", ""}
MAX_FIELD_LENGTH = 1500


def sanitize_value(val):
    """Sanitiza un valor de entrada."""
    if val is None:
        return ""
    s = str(val).strip()
    s = re.sub(r'[\x00-\x08\x0b\x0c\x0e-\x1f]', '', s)
    if len(s) > MAX_FIELD_LENGTH:
        s = s[:MAX_FIELD_LENGTH]
    return s


def validate_registro(data: dict) -> list[str]:
    """Valida los datos del registro. Retorna lista de errores."""
    errors = []

    if not data.get("registro_id"):
        errors.append("registro_id es requerido")

    clasif = data.get("clasificacion_caso", "")
    if clasif and clasif not in VALID_CLASIFICACIONES:
        errors.append(f"clasificacion_caso inválida: {clasif}")

    sexo = data.get("sexo", "")
    if sexo and sexo not in VALID_SEXO:
        errors.append(f"sexo inválido: {sexo}")

    semana = data.get("semana_epidemiologica", "")
    if semana:
        try:
            s = int(semana)
            if s < 1 or s > 53:
                errors.append(f"semana_epidemiologica fuera de rango: {s}")
        except ValueError:
            errors.append(f"semana_epidemiologica no es número: {semana}")

    for field in ("edad_anios", "edad_meses"):
        val = data.get(field, "")
        if val:
            try:
                n = int(val)
                if n < 0 or n > 150:
                    errors.append(f"{field} fuera de rango: {n}")
            except ValueError:
                errors.append(f"{field} no es número: {val}")

    return errors


def validate_registro_update(data: dict) -> list[str]:
    """Valida campos para actualización parcial (todos opcionales)."""
    errors = []

    clasif = data.get("clasificacion_caso", "")
    if clasif and clasif not in VALID_CLASIFICACIONES:
        errors.append(f"clasificacion_caso inválida: {clasif}")

    sexo = data.get("sexo", "")
    if sexo and sexo not in VALID_SEXO:
        errors.append(f"sexo inválido: {sexo}")

    semana = data.get("semana_epidemiologica", "")
    if semana:
        try:
            s = int(semana)
            if s < 1 or s > 53:
                errors.append(f"semana_epidemiologica fuera de rango: {s}")
        except ValueError:
            errors.append(f"semana_epidemiologica no es número: {semana}")

    for field in ("edad_anios", "edad_meses"):
        val = data.get(field, "")
        if val:
            try:
                n = int(val)
                if n < 0 or n > 150:
                    errors.append(f"{field} fuera de rango: {n}")
            except ValueError:
                errors.append(f"{field} no es número: {val}")

    return errors


# ─── Init DB ──────────────────────────────────────────────
@app.on_event("startup")
def startup():
    init_db()
    init_audit_table()
    init_mspas_tables()
    init_godata_tables()
    # Auto-recover records stuck in 'enviando' state (e.g. from crashed processes)
    recovered = recover_stuck_submissions()
    if recovered:
        logger.info(f"Recovered {recovered} stuck MSPAS submissions")
    recovered_gd = recover_stuck_syncs()
    if recovered_gd:
        logger.info(f"Recovered {recovered_gd} stuck GoData syncs")
    print(f"Vigilancia Sarampión API v2.0 iniciada en puerto {PORT}")


# ─── Helpers ──────────────────────────────────────────────
def verify_api_key(x_api_key: str = Header(None)):
    """Verifica la API key con comparación timing-safe."""
    if not x_api_key or not hmac.compare_digest(x_api_key, API_SECRET_KEY):
        raise HTTPException(status_code=401, detail="API key inválida")


def get_client_ip(request: Request) -> str:
    forwarded = request.headers.get("x-forwarded-for")
    if forwarded:
        return forwarded.split(",")[0].strip()
    return request.client.host if request.client else "unknown"


# ─── Endpoints ────────────────────────────────────────────

@app.get("/api/health")
def health():
    return {
        "status": "ok",
        "service": "vigilancia-sarampion-igss",
        "version": "2.0.0",
        "timestamp": datetime.now().isoformat(),
        "total_registros": get_count(),
    }


@app.post("/api/registro")
async def crear_registro(request: Request):
    """Recibe datos del formulario y los guarda en la DB."""
    ip = get_client_ip(request)

    _cleanup_rate_limiter()
    now = time.time()
    last = _rate_limiter.get(ip, 0)
    if now - last < RATE_LIMIT_SECONDS:
        raise HTTPException(
            status_code=429,
            detail=f"Debe esperar {RATE_LIMIT_SECONDS} segundos entre envíos",
        )

    try:
        data = await request.json()
    except Exception:
        raise HTTPException(status_code=400, detail="JSON inválido")

    validation_errors = validate_registro(data)
    if validation_errors:
        raise HTTPException(status_code=400, detail="; ".join(validation_errors))

    afiliacion = sanitize_value(data.get("afiliacion", ""))
    fecha = sanitize_value(data.get("fecha_notificacion", ""))
    if afiliacion and fecha and check_duplicate(afiliacion, fecha):
        raise HTTPException(
            status_code=409,
            detail=f"Ya existe un registro para afiliación {afiliacion} con fecha {fecha}",
        )

    # Período sarampión (21 días): rechazar si ya existe un registro del mismo
    # afiliado cuya fecha de inicio de síntomas (o notificación si falta) está
    # dentro de ±21 días de la entregada. Evita que el mismo caso se notifique
    # dos veces cuando el paciente vuelve días después.
    fis = sanitize_value(data.get("fecha_inicio_sintomas", ""))
    if afiliacion:
        dup21 = find_duplicate_21_dias(
            numero_afiliado=afiliacion,
            fecha_inicio_sintomas=fis,
            fecha_notificacion=fecha,
        )
        if dup21:
            raise HTTPException(
                status_code=409,
                detail=(
                    f"Período sarampión (21 días): ya existe el registro "
                    f"{dup21.get('registro_id')} para este afiliado con fecha "
                    f"cercana ({dup21.get('fecha_inicio_sintomas') or dup21.get('fecha_notificacion')}). "
                    f"Si corresponde a un caso distinto, ajuste la fecha de inicio de síntomas."
                ),
            )

    # Validate that at least a name is provided
    _nombres = sanitize_value(data.get("nombres", ""))
    _apellidos = sanitize_value(data.get("apellidos", ""))
    _nombre_apellido = sanitize_value(data.get("nombre_apellido", ""))
    if not _nombres and not _apellidos and not _nombre_apellido:
        raise HTTPException(400, "Se requiere al menos nombre o apellido del paciente")

    clean = {}
    for col in COLUMNS:
        clean[col] = sanitize_value(data.get(col, ""))

    try:
        registro_id = insert_registro(clean, ip=ip)
    except Exception:
        raise HTTPException(status_code=500, detail="Error al guardar el registro")

    _rate_limiter[ip] = now

    return {
        "success": True,
        "registro_id": registro_id,
        "message": "Registro guardado correctamente",
    }


# ─── Nuevos endpoints: Registro individual ────────────────

@app.get("/api/registro/{registro_id}")
def obtener_registro(registro_id: str, x_api_key: str = Header(None)):
    """Obtiene un registro individual por su ID. Requiere API key."""
    verify_api_key(x_api_key)

    reg = get_registro_by_id(registro_id)
    if not reg:
        raise HTTPException(status_code=404, detail=f"Registro {registro_id} no encontrado")

    return {"data": reg}


@app.put("/api/registro/{registro_id}")
async def actualizar_registro(registro_id: str, request: Request, x_api_key: str = Header(None)):
    """Edita un registro existente. Requiere API key.
    Solo enviar los campos que cambiaron.
    """
    verify_api_key(x_api_key)

    # Verificar que existe
    existing = get_registro_by_id(registro_id)
    if not existing:
        raise HTTPException(status_code=404, detail=f"Registro {registro_id} no encontrado")

    try:
        data = await request.json()
    except Exception:
        raise HTTPException(status_code=400, detail="JSON inválido")

    if not data:
        raise HTTPException(status_code=400, detail="No se proporcionaron campos para actualizar")

    # Validar campos
    validation_errors = validate_registro_update(data)
    if validation_errors:
        raise HTTPException(status_code=400, detail="; ".join(validation_errors))

    # Sanitizar solo campos editables
    clean = {}
    for col in data:
        if col in EDITABLE_COLUMNS:
            clean[col] = sanitize_value(data[col])

    if not clean:
        raise HTTPException(status_code=400, detail="Ningún campo editable proporcionado")

    # Audit trail
    ip = get_client_ip(request)
    log_changes(registro_id, existing, clean, usuario="portal", ip=ip)

    # Actualizar
    success = update_registro(registro_id, clean)
    if not success:
        raise HTTPException(status_code=500, detail="Error actualizando registro")

    return {
        "success": True,
        "registro_id": registro_id,
        "fields_updated": list(clean.keys()),
        "message": "Registro actualizado correctamente",
    }


@app.get("/api/registro/{registro_id}/audit")
def obtener_audit_trail(registro_id: str, x_api_key: str = Header(None)):
    """Obtiene el historial de cambios de un registro. Requiere API key."""
    verify_api_key(x_api_key)

    reg = get_registro_by_id(registro_id)
    if not reg:
        raise HTTPException(status_code=404, detail=f"Registro {registro_id} no encontrado")

    trail = get_audit_trail(registro_id)
    return {"registro_id": registro_id, "total": len(trail), "data": trail}


@app.delete("/api/registro/{registro_id}")
def eliminar_registro(registro_id: str, request: Request, x_api_key: str = Header(None)):
    """Elimina un registro. Guarda snapshot en audit_log para restauración. Requiere API key."""
    verify_api_key(x_api_key)

    reg = get_registro_by_id(registro_id)
    if not reg:
        raise HTTPException(status_code=404, detail=f"Registro {registro_id} no encontrado")

    ip = get_client_ip(request)
    success = delete_registro(registro_id, usuario="portal", ip=ip)
    if not success:
        raise HTTPException(status_code=500, detail="Error eliminando registro")

    return {"success": True, "registro_id": registro_id, "message": "Registro eliminado (snapshot guardado en auditoría)"}


# ─── Nuevo endpoint: Carga masiva ─────────────────────────

@app.post("/api/registros/upload")
async def upload_registros(request: Request, x_api_key: str = Header(None)):
    """Carga masiva de registros desde archivo Excel (.xlsx).
    Requiere API key. Enviar como multipart/form-data con campo 'file'.
    """
    verify_api_key(x_api_key)

    form = await request.form()
    file = form.get("file")
    if not file:
        raise HTTPException(status_code=400, detail="No se proporcionó archivo. Use campo 'file'.")

    # Leer contenido
    contents = await file.read()
    max_bytes = MAX_UPLOAD_SIZE_MB * 1024 * 1024
    if len(contents) > max_bytes:
        raise HTTPException(
            status_code=413,
            detail=f"Archivo muy grande ({len(contents) // 1024 // 1024}MB). Máximo: {MAX_UPLOAD_SIZE_MB}MB",
        )

    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(contents), read_only=True, data_only=True)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Error leyendo archivo Excel: {e}")

    ws = wb.active
    if ws is None:
        raise HTTPException(status_code=400, detail="El archivo Excel no tiene hojas")

    # Leer headers de la primera fila
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not header_row:
        raise HTTPException(status_code=400, detail="El archivo Excel no tiene encabezados")

    headers = [str(h or "").strip().lower().replace(" ", "_") for h in header_row]

    # Mapeo flexible: header Excel → columna DB
    # Acepta tanto nombres internos (afiliacion) como labels (Afiliación)
    HEADER_ALIASES = {
        "no._registro": "registro_id",
        "no_registro": "registro_id",
        "fecha/hora_envío": "timestamp_envio",
        "fecha_hora_envio": "timestamp_envio",
        "diagnóstico_registrado": "diagnostico_registrado",
        "código_cie-10": "codigo_cie10",
        "codigo_cie-10": "codigo_cie10",
        "código_cie10": "codigo_cie10",
        "fecha_notificación": "fecha_notificacion",
        "semana_epidemiológica": "semana_epidemiologica",
        "enviaron_ficha": "envio_ficha",
        "afiliación": "afiliacion",
        "número_de_afiliación": "afiliacion",
        "nombre_y_apellido": "nombre_apellido",
        "edad_(años)": "edad_anios",
        "edad_(meses)": "edad_meses",
        "departamento": "departamento_residencia",
        "municipio": "municipio_residencia",
        "dirección": "direccion_exacta",
        "embarazada": "esta_embarazada",
        "fiebre": "signo_fiebre",
        "exantema": "signo_exantema",
        "manchas_koplik": "signo_manchas_koplik",
        "tos": "signo_tos",
        "conjuntivitis": "signo_conjuntivitis",
        "artralgia": "signo_artralgia",
        "coriza": "signo_coriza",
        "adenopatías": "signo_adenopatias",
        "adenopatias": "signo_adenopatias",
        "dosis_spr/sr": "numero_dosis_spr",
        "dosis_spr_sr": "numero_dosis_spr",
        "fecha_última_dosis": "fecha_ultima_dosis",
        "condición_egreso": "condicion_egreso",
        "fecha_defunción": "fecha_defuncion",
        "igg_cualitativo": "resultado_igg_cualitativo",
        "igm_cualitativo": "resultado_igm_cualitativo",
        "rt-pcr_orina": "resultado_pcr_orina",
        "rt-pcr_hisopado": "resultado_pcr_hisopado",
        "clasificación": "clasificacion_caso",
        "clasificacion": "clasificacion_caso",
    }

    # Resolver mapeo de columnas
    col_map = {}  # index -> db_column
    for i, h in enumerate(headers):
        if h in COLUMNS:
            col_map[i] = h
        elif h in HEADER_ALIASES:
            col_map[i] = HEADER_ALIASES[h]

    if not col_map:
        raise HTTPException(
            status_code=400,
            detail=f"No se reconocieron columnas. Headers encontrados: {headers[:10]}",
        )

    # Leer filas de datos
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_data = {}
        for i, val in enumerate(row):
            if i in col_map:
                row_data[col_map[i]] = sanitize_value(val)
        # Skip filas completamente vacías
        if any(v for v in row_data.values()):
            # Generar registro_id si no viene
            if not row_data.get("registro_id"):
                from datetime import datetime as dt
                import random
                import string
                ts = int((dt.now() - dt(dt.now().year, 1, 1)).total_seconds())
                code = ''.join(random.choices(
                    [c for c in string.ascii_uppercase + string.digits if c not in 'IO01'], k=4
                ))
                row_data["registro_id"] = f"IGSS-SAR-{dt.now().year}-{ts:07d}-{code}"
            if not row_data.get("timestamp_envio"):
                row_data["timestamp_envio"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            rows.append(row_data)

    wb.close()

    if not rows:
        raise HTTPException(status_code=400, detail="El archivo no contiene filas de datos")

    # Insertar
    ip = get_client_ip(request)
    result = bulk_insert_registros(rows, ip=ip)

    return {
        "success": True,
        "total_rows": len(rows),
        "inserted": result["inserted"],
        "duplicates": result["duplicates"],
        "errors": result["errors"][:50],
        "columns_mapped": {v: headers[k] for k, v in col_map.items()},
        "message": f"Carga completada: {result['inserted']} insertados, {result['duplicates']} duplicados",
    }


# ─── Endpoints existentes ─────────────────────────────────

@app.get("/api/registros")
def listar_registros(
    x_api_key: str = Header(None),
    limit: int = Query(100, ge=1, le=5000),
    offset: int = Query(0, ge=0),
    clasificacion: str = Query(None),
    departamento: str = Query(None),
    semana: str = Query(None),
):
    """Lista registros con paginación y filtros. Requiere API key."""
    verify_api_key(x_api_key)

    filters = {}
    if clasificacion:
        filters["clasificacion_caso"] = clasificacion
    if departamento:
        filters["departamento_residencia"] = departamento
    if semana:
        filters["semana_epidemiologica"] = semana

    registros = get_registros(limit=limit, offset=offset, filters=filters)
    total = get_count(filters=filters)

    return {
        "total": total,
        "limit": limit,
        "offset": offset,
        "data": registros,
    }


@app.get("/api/registros/search")
def api_search_registros(
    q: str = Query("", description="Texto a buscar"),
    limit: int = Query(20, ge=1, le=100),
    x_api_key: str = Header(None),
):
    """Buscar registros por afiliación, nombre o ID."""
    verify_api_key(x_api_key)
    if not q or len(q) < 2:
        raise HTTPException(400, "La búsqueda debe tener al menos 2 caracteres")
    results = search_registros(q, limit)
    return {"data": results, "total": len(results), "query": q}


@app.get("/api/registros/count")
def contar_registros(
    x_api_key: str = Header(None),
    clasificacion: str = Query(None),
):
    """Cuenta total de registros. Requiere API key."""
    verify_api_key(x_api_key)

    filters = {}
    if clasificacion:
        filters["clasificacion_caso"] = clasificacion

    return {"total": get_count(filters=filters)}


@app.get("/api/export/excel")
def export_excel(x_api_key: str = Header(None)):
    """Exporta registros como Excel con formato IGSS oficial.
    3 hojas (SOSPECHOSOS, CONFIRMADOS, SUSPENDIDOS), 2 filas header (categorías + columnas).
    """
    verify_api_key(x_api_key)

    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    registros = get_registros(limit=50000)

    # Separar por clasificación
    sheets_data = {"SOSPECHOSOS": [], "CONFIRMADOS": [], "SUSPENDIDOS": []}
    for r in registros:
        clasif = (r.get("clasificacion_caso") or "").upper()
        if clasif in ("CONFIRMADO",):
            sheets_data["CONFIRMADOS"].append(r)
        elif clasif in ("SUSPENDIDO",):
            sheets_data["SUSPENDIDOS"].append(r)
        else:
            sheets_data["SOSPECHOSOS"].append(r)

    # Columnas en orden del Excel — compatible MSPAS + IGSS
    EXPORT_COLS = [
        # Datos Generales (1-13)
        "diagnostico_registrado", "codigo_cie10", "unidad_medica", "unidad_medica_otra",
        "centro_externo",
        "fecha_registro_diagnostico", "fecha_notificacion", "semana_epidemiologica",
        "servicio_reporta", "nom_responsable", "cargo_responsable", "telefono_responsable",
        "envio_ficha",
        # Datos del Paciente (14-31)
        "afiliacion", "nombres", "apellidos", "sexo", "fecha_nacimiento",
        "edad_anios", "edad_meses", "edad_dias", "pueblo_etnia", "comunidad_linguistica",
        "ocupacion", "escolaridad",
        "departamento_residencia", "municipio_residencia", "poblado",
        "direccion_exacta", "nombre_encargado", "telefono_encargado",
        # Embarazo (32-37)
        "esta_embarazada", "lactando", "semanas_embarazo", "fecha_probable_parto",
        "vacuna_embarazada", "fecha_vacunacion_embarazada",
        # Información Clínica (38-65)
        "fecha_inicio_sintomas", "fecha_captacion", "fuente_notificacion",
        "fuente_notificacion_otra",
        "fecha_visita_domiciliaria", "fecha_inicio_investigacion", "busqueda_activa",
        "busqueda_activa_otra",
        "fecha_inicio_erupcion", "sitio_inicio_erupcion", "sitio_inicio_erupcion_otro",
        "fecha_inicio_fiebre", "temperatura_celsius",
        "signo_fiebre", "signo_exantema", "signo_manchas_koplik",
        "signo_tos", "signo_conjuntivitis", "signo_artralgia",
        "signo_coriza", "signo_adenopatias", "asintomatico",
        "vacunado", "fuente_info_vacuna", "tipo_vacuna", "numero_dosis_spr",
        "fecha_ultima_dosis", "observaciones_vacuna",
        # Hospitalización (66-77)
        "hospitalizado", "hosp_nombre", "hosp_fecha", "no_registro_medico",
        "complicaciones", "complicaciones_otra", "diagnostico_otro",
        "condicion_egreso", "fecha_egreso", "fecha_defuncion",
        "medico_certifica_defuncion", "motivo_consulta",
        # Factores de Riesgo (78-83)
        "contacto_sospechoso_7_23", "caso_sospechoso_comunidad_3m",
        "viajo_7_23_previo", "destino_viaje",
        "contacto_enfermo_catarro", "contacto_embarazada",
        # Laboratorio (84-105)
        "recolecto_muestra", "motivo_no_recoleccion", "muestra_suero", "muestra_suero_fecha",
        "muestra_hisopado", "muestra_hisopado_fecha",
        "muestra_orina", "muestra_orina_fecha",
        "muestra_otra", "muestra_otra_descripcion", "muestra_otra_fecha",
        "antigeno_prueba", "antigeno_otro", "resultado_prueba",
        "resultado_igg_cualitativo", "resultado_igm_cualitativo",
        "resultado_pcr_orina", "resultado_pcr_hisopado",
        "fecha_recepcion_laboratorio", "fecha_resultado_laboratorio",
        "resultado_igg_numerico", "resultado_igm_numerico",
        # Contactos y IGSS (106-121)
        "contactos_directos", "clasificacion_caso", "fecha_clasificacion_final",
        "responsable_clasificacion", "observaciones",
        "es_empleado_igss", "unidad_medica_trabaja", "puesto_desempena",
        "subgerencia_igss", "subgerencia_igss_otra",
        "direccion_igss", "direccion_igss_otra",
        "departamento_igss", "departamento_igss_otro",
        "seccion_igss", "seccion_igss_otra",
        # Formato 2026 — Encabezado (122-123)
        "diagnostico_sospecha", "diagnostico_sospecha_otro",
        # Formato 2026 — Unidad (124-130)
        "area_salud_mspas", "distrito_salud_mspas", "servicio_salud_mspas", "correo_responsable",
        "es_seguro_social", "establecimiento_privado", "establecimiento_privado_nombre",
        # Formato 2026 — Paciente (131-139)
        "tipo_identificacion", "numero_identificacion", "parentesco_tutor", "tipo_id_tutor", "numero_id_tutor",
        "es_migrante", "trimestre_embarazo", "telefono_paciente", "pais_residencia",
        # Formato 2026 — Antecedentes (140-152)
        "tiene_antecedentes_medicos", "antecedentes_medicos_detalle",
        "antecedente_desnutricion", "antecedente_inmunocompromiso", "antecedente_enfermedad_cronica",
        "dosis_spr", "fecha_ultima_spr", "dosis_sr", "fecha_ultima_sr", "dosis_sprv", "fecha_ultima_sprv",
        "sector_vacunacion",
        # Formato 2026 — Clínica (153-162)
        "tiene_complicaciones",
        "comp_neumonia", "comp_encefalitis", "comp_diarrea", "comp_trombocitopenia",
        "comp_otitis", "comp_ceguera", "comp_otra_texto",
        "aislamiento_respiratorio", "fecha_aislamiento",
        # Formato 2026 — Factores de Riesgo (163-171)
        "viaje_pais", "viaje_departamento", "viaje_municipio", "viaje_fecha_salida", "viaje_fecha_entrada",
        "familiar_viajo_exterior", "familiar_fecha_retorno", "fuente_posible_contagio", "fuente_contagio_otro",
        # Formato 2026 — Acciones de Respuesta (172-180)
        "bai_realizada", "bai_casos_sospechosos", "bac_realizada", "bac_casos_sospechosos",
        "vacunacion_bloqueo", "monitoreo_rapido_vacunacion", "vacunacion_barrido",
        "vitamina_a_administrada", "vitamina_a_dosis",
        # Formato 2026 — Laboratorio adicional (181-183)
        "motivo_no_3_muestras", "secuenciacion_resultado", "secuenciacion_fecha",
        # Formato 2026 — Clasificación (184-193)
        "criterio_confirmacion", "contacto_otro_caso", "contacto_otro_caso_detalle",
        "criterio_descarte", "fuente_infeccion", "pais_importacion",
        "caso_analizado_por", "caso_analizado_por_otro", "condicion_final_paciente", "causa_muerte_certificado",
        # GoData (194-197)
        "godata_case_id", "godata_sync_status", "godata_last_sync", "form_version",
    ]

    COL_HEADERS = [
        # Datos Generales
        "Diagnóstico", "CIE-10", "Unidad Médica", "Unidad Médica Otra",
        "Centro Externo",
        "Fecha Registro", "Fecha Notificación", "Semana Epi.",
        "Servicio", "Responsable", "Cargo", "Teléfono Resp.",
        "Enviaron Ficha",
        # Datos del Paciente
        "Afiliación", "Nombres", "Apellidos", "Sexo", "Fecha Nac.",
        "Edad Años", "Edad Meses", "Edad Días", "Pueblo/Etnia", "Comunidad Ling.",
        "Ocupación", "Escolaridad",
        "Departamento", "Municipio", "Poblado",
        "Dirección", "Encargado", "Tel. Encargado",
        # Embarazo
        "Embarazada", "Lactando", "Sem. Embarazo", "Fecha Prob. Parto",
        "Vacuna Emb.", "Fecha Vac. Emb.",
        # Información Clínica
        "Fecha Inicio Síntomas", "Fecha Captación", "Fuente Notif.",
        "Fuente Notif. Otra",
        "Fecha Visita Dom.", "Fecha Inicio Invest.", "Búsq. Activa",
        "Búsq. Activa Otra",
        "Fecha Inicio Erupción", "Sitio Erupción", "Otro Sitio Erupción",
        "Fecha Inicio Fiebre", "Temp. °C",
        "Fiebre", "Exantema", "Koplik",
        "Tos", "Conjuntivitis", "Artralgia",
        "Coriza", "Adenopatías", "Asintomático",
        "Vacunado", "Fuente Info Vacuna", "Tipo Vacuna", "No. Dosis",
        "Fecha Últ. Dosis", "Obs. Vacunación",
        # Hospitalización
        "Hospitalizado", "Hospital", "Fecha Hosp.", "Reg. Médico",
        "Complicaciones", "Complicaciones (otra)", "Diagnóstico (otro)",
        "Condición Egreso", "Fecha Egreso", "Fecha Defunción",
        "Médico Defunción", "Motivo Consulta",
        # Factores de Riesgo
        "Contacto Sosp. 7-23d", "Caso Sosp. Comunidad 3m",
        "Viajó 7-23d", "Destino Viaje",
        "Contacto Enfermo", "Contacto Embarazada",
        # Laboratorio
        "Recolectó Muestra", "Motivo No Recolección", "Suero", "Fecha Suero",
        "Hisopado", "Fecha Hisopado",
        "Orina", "Fecha Orina",
        "Otra Muestra", "Desc. Otra Muestra", "Fecha Otra Muestra",
        "Antígeno", "Otro Antígeno", "Resultado Prueba",
        "IgG Cual.", "IgM Cual.",
        "PCR Orina", "PCR Hisopado",
        "Fecha Recep. Lab", "Fecha Result. Lab",
        "IgG Num.", "IgM Num.",
        # Contactos y IGSS
        "Contactos Directos", "Clasificación", "Fecha Clasif. Final",
        "Resp. Clasificación", "Observaciones",
        "Empleado IGSS", "Unidad Trabaja", "Puesto",
        "Subgerencia IGSS", "Subgerencia Otra",
        "Dirección IGSS", "Dirección Otra",
        "Depto. IGSS", "Depto. Otro",
        "Sección IGSS", "Sección Otra",
        # Formato 2026 — Encabezado
        "Diagnóstico Sospecha", "Diagnóstico Sospecha (otro)",
        # Formato 2026 — Unidad
        "Área Salud MSPAS", "Distrito Salud MSPAS", "Servicio Salud MSPAS", "Correo Responsable",
        "¿Seguro Social?", "¿Estab. Privado?", "Nombre Estab. Privado",
        # Formato 2026 — Paciente
        "Tipo Identificación", "No. Identificación", "Parentesco Tutor", "Tipo ID Tutor", "No. ID Tutor",
        "¿Migrante?", "Trimestre Embarazo", "Teléfono Paciente", "País Residencia",
        # Formato 2026 — Antecedentes
        "¿Antecedentes Médicos?", "Detalle Antecedentes",
        "Desnutrición", "Inmunocompromiso", "Enf. Crónica",
        "Dosis SPR", "Fecha Últ. SPR", "Dosis SR", "Fecha Últ. SR", "Dosis SPRv", "Fecha Últ. SPRv",
        "Sector Vacunación",
        # Formato 2026 — Clínica
        "¿Tiene Complicaciones?",
        "Comp. Neumonía", "Comp. Encefalitis", "Comp. Diarrea", "Comp. Trombocitopenia",
        "Comp. Otitis", "Comp. Ceguera", "Comp. Otra (texto)",
        "Aislamiento Respiratorio", "Fecha Aislamiento",
        # Formato 2026 — Factores de Riesgo
        "País Viaje", "Depto. Viaje", "Municipio Viaje", "Fecha Salida Viaje", "Fecha Entrada Viaje",
        "¿Familiar Viajó?", "Fecha Retorno Familiar", "Fuente Posible Contagio", "Fuente Contagio (otro)",
        # Formato 2026 — Acciones de Respuesta
        "BAI Realizada", "BAI Casos Sosp.", "BAC Realizada", "BAC Casos Sosp.",
        "Vac. Bloqueo", "Monitoreo Rápido Vac.", "Vac. Barrido",
        "Vitamina A", "Dosis Vitamina A",
        # Formato 2026 — Laboratorio adicional
        "Motivo No 3 Muestras", "Result. Secuenciación", "Fecha Secuenciación",
        # Formato 2026 — Clasificación
        "Criterio Confirmación", "Contacto Otro Caso", "Detalle Contacto",
        "Criterio Descarte", "Fuente Infección", "País Importación",
        "Analizado Por", "Analizado Por (otro)", "Condición Final", "Causa Muerte (certificado)",
        # GoData
        "GoData Case ID", "GoData Sync Status", "GoData Últ. Sync", "Versión Formulario",
    ]

    # Category headers (row 1) — (start_col, end_col, label)  1-indexed within EXPORT_COLS
    CATEGORIES = [
        (1, 13, "DATOS GENERALES"),
        (14, 31, "DATOS DEL PACIENTE"),
        (32, 37, "EMBARAZO"),
        (38, 65, "INFORMACIÓN CLÍNICA"),
        (66, 77, "HOSPITALIZACIÓN"),
        (78, 83, "FACTORES DE RIESGO"),
        (84, 105, "LABORATORIO"),
        (106, 121, "CONTACTOS Y DATOS IGSS"),
        # Formato 2026
        (122, 123, "FORMATO 2026 — ENCABEZADO"),
        (124, 130, "FORMATO 2026 — UNIDAD"),
        (131, 139, "FORMATO 2026 — PACIENTE"),
        (140, 151, "FORMATO 2026 — ANTECEDENTES"),
        (152, 161, "FORMATO 2026 — CLÍNICA"),
        (162, 170, "FORMATO 2026 — FACTORES DE RIESGO"),
        (171, 179, "FORMATO 2026 — ACCIONES DE RESPUESTA"),
        (180, 182, "FORMATO 2026 — LABORATORIO"),
        (183, 192, "FORMATO 2026 — CLASIFICACIÓN"),
        (193, 196, "GODATA"),
    ]

    # Styles
    cat_fill = PatternFill(start_color="1B5E20", end_color="1B5E20", fill_type="solid")
    cat_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    hdr_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
    hdr_font = Font(name="Calibri", bold=True, color="FFFFFF", size=9)
    border_thin = Border(
        bottom=Side(style="thin"), right=Side(style="thin", color="CCCCCC"),
    )
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    wb = openpyxl.Workbook()
    first = True

    for sheet_name, data in sheets_data.items():
        if first:
            ws = wb.active
            ws.title = sheet_name
            first = False
        else:
            ws = wb.create_sheet(sheet_name)

        # Add "No. caso" as first column
        total_cols = len(EXPORT_COLS) + 1  # +1 for No. caso

        # Row 1: Category headers (merged)
        ws.cell(row=1, column=1, value="No.").fill = cat_fill
        ws.cell(row=1, column=1).font = cat_font
        ws.cell(row=1, column=1).alignment = center_align

        for start, end, label in CATEGORIES:
            s = start + 1  # offset for No. caso column
            e = end + 1
            cell = ws.cell(row=1, column=s, value=label)
            cell.fill = cat_fill
            cell.font = cat_font
            cell.alignment = center_align
            if s != e:
                ws.merge_cells(start_row=1, start_column=s, end_row=1, end_column=e)

        # Row 2: Column headers
        ws.cell(row=2, column=1, value="No. caso").fill = hdr_fill
        ws.cell(row=2, column=1).font = hdr_font
        ws.cell(row=2, column=1).alignment = center_align
        ws.cell(row=2, column=1).border = border_thin

        for i, header in enumerate(COL_HEADERS):
            cell = ws.cell(row=2, column=i + 2, value=header)
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.alignment = center_align
            cell.border = border_thin

        # Data rows
        for row_idx, reg in enumerate(data, 3):
            ws.cell(row=row_idx, column=1, value=row_idx - 2)  # No. caso
            for col_idx, col_key in enumerate(EXPORT_COLS, 2):
                ws.cell(row=row_idx, column=col_idx, value=reg.get(col_key, ""))

        # Column widths
        ws.column_dimensions["A"].width = 6
        for i, col_key in enumerate(EXPORT_COLS, 2):
            letter = get_column_letter(i)
            ws.column_dimensions[letter].width = min(max(len(COL_HEADERS[i-2]), 10) + 2, 28)

        ws.freeze_panes = "A3"  # Freeze first 2 rows

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"BASE_DATOS_VIGILANCIA_SARAMPION_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@app.get("/api/export/csv")
def export_csv(x_api_key: str = Header(None)):
    """Exporta todos los registros como CSV. Requiere API key."""
    verify_api_key(x_api_key)

    registros = get_registros(limit=50000)
    export_cols = [c for c in COLUMNS if c not in ("ip_origen", "created_at")]

    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=export_cols, extrasaction="ignore")
    writer.writeheader()
    for reg in registros:
        writer.writerow({k: reg.get(k, "") for k in export_cols})

    output.seek(0)
    filename = f"vigilancia_sarampion_igss_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
    return StreamingResponse(
        io.BytesIO(output.getvalue().encode("utf-8-sig")),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


# ─── PDF Ficha Export ─────────────────────────────────────
@app.get("/api/export/ficha/{registro_id}")
def export_ficha_pdf(registro_id: str, x_api_key: str = Header(None)):
    """Genera PDF de ficha epidemiológica para un registro individual."""
    verify_api_key(x_api_key)
    reg = get_registro_by_id(registro_id)
    if not reg:
        raise HTTPException(status_code=404, detail=f"Registro {registro_id} no encontrado")

    from pdf_ficha import generar_ficha_pdf
    pdf_bytes = generar_ficha_pdf(reg)

    filename = f"ficha_{registro_id.replace('/', '_')}.pdf"
    return StreamingResponse(
        io.BytesIO(pdf_bytes),
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@app.get("/api/export/ficha-v2/{registro_id}")
def export_ficha_v2(registro_id: str, x_api_key: str = Header(None)):
    """Genera PDF ficha epidemiológica formato MSPAS 2026."""
    verify_api_key(x_api_key)
    from database import get_registro_by_id
    record = get_registro_by_id(registro_id)
    if not record:
        raise HTTPException(404, "Registro no encontrado")
    from pdf_ficha_v2 import generar_ficha_v2_pdf
    pdf_bytes = generar_ficha_v2_pdf(record)
    return StreamingResponse(
        io.BytesIO(pdf_bytes),
        media_type="application/pdf",
        headers={"Content-Disposition": f"inline; filename=ficha_v2_{registro_id}.pdf"}
    )


@app.post("/api/export/fichas")
async def export_fichas_bulk(request: Request, x_api_key: str = Header(None)):
    """Genera PDFs para múltiples registros.
    Body: {"ids": ["IGSS-SAR-...", ...], "format": "merged"|"zip"}
    """
    verify_api_key(x_api_key)
    data = await request.json()
    ids = data.get("ids", [])
    fmt = data.get("format", "merged")

    if not ids or len(ids) > 500:
        raise HTTPException(status_code=400, detail="Proporcione entre 1 y 500 IDs")

    records = []
    for rid in ids:
        reg = get_registro_by_id(rid)
        if reg:
            records.append(reg)

    if not records:
        raise HTTPException(status_code=404, detail="Ningún registro encontrado")

    from pdf_ficha import generar_fichas_bulk
    merge = (fmt == "merged")
    result_bytes = generar_fichas_bulk(records, merge=merge)

    if merge:
        filename = f"fichas_sarampion_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        return StreamingResponse(
            io.BytesIO(result_bytes),
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )
    else:
        filename = f"fichas_sarampion_{datetime.now().strftime('%Y%m%d_%H%M%S')}.zip"
        return StreamingResponse(
            io.BytesIO(result_bytes),
            media_type="application/zip",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )


@app.post("/api/export/fichas-v2")
async def export_fichas_v2_bulk(request: Request, x_api_key: str = Header(None)):
    """Genera PDFs formato MSPAS 2026 (GoData) para múltiples registros.
    Body: {"ids": ["IGSS-SAR-...", ...], "format": "merged"|"zip"}
    """
    verify_api_key(x_api_key)
    data = await request.json()
    ids = data.get("ids", [])
    fmt = data.get("format", "merged")

    if not ids or len(ids) > 500:
        raise HTTPException(status_code=400, detail="Proporcione entre 1 y 500 IDs")

    records = []
    for rid in ids:
        reg = get_registro_by_id(rid)
        if reg:
            records.append(reg)

    if not records:
        raise HTTPException(status_code=404, detail="Ningún registro encontrado")

    from pdf_ficha_v2 import generar_fichas_v2_bulk
    merge = (fmt == "merged")
    result_bytes = generar_fichas_v2_bulk(records, merge=merge)

    if merge:
        filename = f"fichas_godata_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        return StreamingResponse(
            io.BytesIO(result_bytes),
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )
    else:
        filename = f"fichas_godata_{datetime.now().strftime('%Y%m%d_%H%M%S')}.zip"
        return StreamingResponse(
            io.BytesIO(result_bytes),
            media_type="application/zip",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )


@app.get("/api/ficha-publica/{registro_id}")
def export_ficha_publica(registro_id: str):
    """Public endpoint for downloading ficha PDF (only for recent records, <30 min old)."""
    reg = get_registro_by_id(registro_id)
    if not reg:
        raise HTTPException(status_code=404, detail="Registro no encontrado")

    # Security: only allow download of records created in the last 30 minutes
    created = reg.get("created_at", "")
    if created:
        try:
            created_dt = datetime.fromisoformat(created)
            age_minutes = (datetime.now() - created_dt).total_seconds() / 60
            if age_minutes > 30:
                raise HTTPException(status_code=403, detail="El PDF solo está disponible dentro de los 30 minutos posteriores al envío")
        except (ValueError, TypeError):
            pass

    # Usar ficha v2 (formato MSPAS 2026/GoData) — EPIWEB deprecado
    from pdf_ficha_v2 import generar_ficha_v2_pdf
    pdf_bytes = generar_ficha_v2_pdf(reg)

    filename = f"ficha_godata_{registro_id.replace('/', '_')}.pdf"
    return StreamingResponse(
        io.BytesIO(pdf_bytes),
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


# ─── MSPAS Queue Endpoints ─────────────────────────────────

@app.post("/api/mspas/config")
async def mspas_save_config(request: Request, x_api_key: str = Header(None)):
    """Save MSPAS EPIWEB credentials."""
    verify_api_key(x_api_key)
    data = await request.json()
    username = data.get("username", "")
    password = data.get("password", "")
    if not username or not password:
        raise HTTPException(400, "Username and password required")
    save_credentials(username, password)
    return {"status": "ok", "message": "Credentials saved"}


@app.get("/api/mspas/config")
def mspas_get_config(x_api_key: str = Header(None)):
    """Check if MSPAS credentials are configured."""
    verify_api_key(x_api_key)
    username, password = get_credentials()
    return {
        "configured": bool(username and password),
        "username": username[:3] + "***" if username else None,
    }


@app.post("/api/mspas/test")
def mspas_test_connection(x_api_key: str = Header(None)):
    """Test MSPAS login without submitting anything."""
    verify_api_key(x_api_key)
    username, password = get_credentials()
    if not username:
        raise HTTPException(400, "MSPAS credentials not configured")

    from mspas_bot import MSPASBot
    bot = MSPASBot(username, password)
    from playwright.sync_api import sync_playwright
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        page = browser.new_page()
        try:
            success = bot.login(page)
            return {"success": success, "message": "Login exitoso" if success else "Login fallido"}
        except Exception as e:
            return {"success": False, "message": str(e)}
        finally:
            browser.close()


@app.get("/api/mspas/queue")
def mspas_get_queue(
    estado: str = Query(None),
    limit: int = Query(50, ge=1, le=500),
    offset: int = Query(0, ge=0),
    x_api_key: str = Header(None),
):
    """Get MSPAS submission queue."""
    verify_api_key(x_api_key)
    items = get_queue(estado=estado, limit=limit, offset=offset)
    counts = get_queue_counts()
    return {"data": items, "counts": counts, "total": sum(counts.values())}


@app.post("/api/mspas/queue/enqueue-all")
def mspas_enqueue_all(x_api_key: str = Header(None)):
    """Add all records not yet in queue."""
    verify_api_key(x_api_key)
    added = enqueue_all_pending()
    return {"added": added}


@app.post("/api/mspas/queue/approve")
async def mspas_approve(request: Request, x_api_key: str = Header(None)):
    """Approve records for MSPAS submission."""
    verify_api_key(x_api_key)
    data = await request.json()
    ids = data.get("ids", [])
    if not ids:
        raise HTTPException(400, "No IDs provided")
    approve_records(ids, aprobado_por="api")
    return {"approved": len(ids)}


@app.post("/api/mspas/submit/{registro_id}")
def mspas_submit_one(registro_id: str, x_api_key: str = Header(None)):
    """Submit a single record to MSPAS (or test-fill in non-production mode)."""
    verify_api_key(x_api_key)

    # Auto-recover any stuck submissions before processing
    recovered = recover_stuck_submissions(timeout_minutes=10)
    if recovered:
        logger.info(f"Auto-recovered {recovered} stuck MSPAS submissions before submit")

    username, password = get_credentials()
    if not username:
        raise HTTPException(400, "MSPAS credentials not configured")

    reg = get_registro_by_id(registro_id)
    if not reg:
        raise HTTPException(404, f"Registro {registro_id} not found")

    if not try_claim_for_submission(registro_id):
        raise HTTPException(409, "Record is already being processed or not in 'aprobado' state")

    from mspas_bot import MSPASBot
    bot = MSPASBot(username, password)
    result = bot.process_record(reg)

    if result.get("duplicate"):
        # Patient already exists in MSPAS — mark as duplicate (confirmed or possible)
        if result.get("duplicate_type") == "possible":
            mark_possible_duplicate(registro_id, result.get("details", ""))
        else:
            mark_duplicate(registro_id, mspas_ficha_id=result.get("mspas_ficha_id", ""))
    elif result.get("success"):
        if result.get("submitted"):
            mark_sent(registro_id, result.get("mspas_ficha_id", ""),
                     result.get("screenshots", [""])[-1] if result.get("screenshots") else "")
        else:
            update_estado(registro_id, 'aprobado')  # Back to approved (test mode)
    else:
        mark_error(registro_id, "; ".join(result.get("errors", ["Unknown error"])))

    return result


@app.post("/api/mspas/submit-batch")
def mspas_submit_batch(x_api_key: str = Header(None)):
    """Submit ALL approved records in batch (single browser session).

    Much faster than individual submissions: 1 login instead of N logins,
    saving ~15s per record. For 1000 records: ~14h instead of ~18h.
    """
    verify_api_key(x_api_key)

    # Recover any stuck submissions first
    recovered = recover_stuck_submissions(timeout_minutes=10)
    if recovered:
        logger.info(f"Auto-recovered {recovered} stuck MSPAS submissions before batch")

    username, password = get_credentials()
    if not username:
        raise HTTPException(400, "MSPAS credentials not configured")

    # Get all approved records
    approved_ids = get_approved_for_submission(limit=99999)
    if not approved_ids:
        return {"processed": 0, "success": 0, "errors": 0,
                "message": "No approved records pending submission"}

    # Claim all for submission (atomically mark as 'enviando')
    claimed = []
    for rid in approved_ids:
        if try_claim_for_submission(rid):
            claimed.append(rid)

    if not claimed:
        return {"processed": 0, "success": 0, "errors": 0,
                "message": "All approved records are already being processed"}

    # Fetch full record data for each claimed ID
    records = []
    for rid in claimed:
        reg = get_registro_by_id(rid)
        if reg:
            records.append(reg)
        else:
            mark_error(rid, "Registro no encontrado en la base de datos")

    if not records:
        return {"processed": 0, "success": 0, "errors": len(claimed),
                "message": "No valid records found for claimed IDs"}

    # Process entire batch with a single browser session.
    # State is updated PER RECORD via on_complete callback so that if the
    # process crashes at record N, records 1..N-1 already have their final
    # state persisted (not stuck in 'enviando').
    from mspas_bot import MSPASBot
    bot = MSPASBot(username, password)

    success_count = 0
    error_count = 0
    duplicate_count = 0

    def _on_record_done(rid, result):
        nonlocal success_count, error_count, duplicate_count
        try:
            if result.get('duplicate'):
                if result.get('duplicate_type') == 'possible':
                    mark_possible_duplicate(rid, result.get('details', ''))
                else:
                    mark_duplicate(rid, mspas_ficha_id=result.get('mspas_ficha_id', ''))
                duplicate_count += 1
            elif result.get('success'):
                if result.get('submitted'):
                    mark_sent(rid, result.get('mspas_ficha_id', ''),
                             result.get('screenshots', [''])[-1] if result.get('screenshots') else '')
                    success_count += 1
                else:
                    # Test mode: back to approved
                    update_estado(rid, 'aprobado')
                    success_count += 1
            else:
                mark_error(rid, '; '.join(result.get('errors', ['Unknown error'])))
                error_count += 1
        except Exception as e:
            logger.error("Failed to update state for record %s: %s", rid, e)
            error_count += 1

    results = bot.process_batch(records, on_complete=_on_record_done)

    return {
        "processed": len(results),
        "success": success_count,
        "errors": error_count,
        "duplicates": duplicate_count,
        "message": f"Batch complete: {success_count} ok, {error_count} errors, {duplicate_count} duplicates",
    }


@app.get("/api/mspas/status/{registro_id}")
def mspas_get_status(registro_id: str, x_api_key: str = Header(None)):
    """Get MSPAS submission status for a record."""
    verify_api_key(x_api_key)
    item = get_status_by_id(registro_id)
    if not item:
        raise HTTPException(404, "Record not in MSPAS queue")
    return item


@app.get("/api/mspas/screenshot/{filename}")
def mspas_get_screenshot(filename: str, x_api_key: str = Header(None)):
    """Serve a bot screenshot file (legacy flat path)."""
    verify_api_key(x_api_key)
    import os
    safe_filename = os.path.basename(filename)
    screenshot_dir = os.environ.get(
        "MSPAS_SCREENSHOT_DIR",
        "/opt/vigilancia-sarampion/data/mspas_screenshots"
    )
    path = os.path.join(screenshot_dir, safe_filename)
    if not os.path.exists(path):
        raise HTTPException(404, "Screenshot not found")
    return FileResponse(path, media_type="image/png")


@app.get("/api/mspas/screenshot/{registro_id}/{filename}")
def mspas_get_screenshot_by_record(registro_id: str, filename: str,
                                    x_api_key: str = Header(None)):
    """Serve a bot screenshot file from a per-record subdirectory."""
    verify_api_key(x_api_key)
    import os
    safe_id = re.sub(r'[^a-zA-Z0-9_-]', '_', registro_id)
    safe_filename = os.path.basename(filename)
    screenshot_dir = os.environ.get(
        "MSPAS_SCREENSHOT_DIR",
        "/opt/vigilancia-sarampion/data/mspas_screenshots"
    )
    path = os.path.join(screenshot_dir, safe_id, safe_filename)
    if not os.path.exists(path):
        raise HTTPException(404, "Screenshot not found")
    return FileResponse(path, media_type="image/png")


# ═══════════════════════════════════════════════════════════
# GoData Endpoints
# ═══════════════════════════════════════════════════════════

@app.post("/api/godata/config")
async def godata_save_config(request: Request, x_api_key: str = Header(None)):
    """Guardar configuración GoData (URL, credenciales, outbreak ID)."""
    verify_api_key(x_api_key)
    body = await request.json()
    godata_url = body.get("godata_url", "").strip().rstrip("/")
    username = body.get("username", "").strip()
    password = body.get("password", "")
    outbreak_id = body.get("outbreak_id", "").strip()
    outbreak_name = body.get("outbreak_name", "").strip()
    if not godata_url or not username or not password or not outbreak_id:
        raise HTTPException(400, "Faltan campos obligatorios: godata_url, username, password, outbreak_id")
    result = save_godata_config(godata_url, username, password, outbreak_id, outbreak_name)
    return result


@app.get("/api/godata/config")
def godata_get_config_endpoint(x_api_key: str = Header(None)):
    """Obtener configuración GoData (sin password)."""
    verify_api_key(x_api_key)
    return get_godata_config()


@app.post("/api/godata/test")
async def godata_test_connection(request: Request, x_api_key: str = Header(None)):
    """Probar conexión con GoData. Accepts temp credentials in body or uses saved config."""
    verify_api_key(x_api_key)
    # Try to read body (may be empty for saved-config test)
    body = {}
    try:
        if request.headers.get("content-type", "").startswith("application/json"):
            body = await request.json()
    except Exception:
        pass
    # Use credentials from body (temporary test) or from saved config
    if body and body.get("godata_url"):
        url = body["godata_url"].strip().rstrip("/")
        user = body.get("username", "").strip()
        pwd = body.get("password", "")
        outbreak_id = body.get("outbreak_id", "").strip()
    else:
        url, user, pwd, outbreak_id = get_godata_credentials()
    if not url:
        raise HTTPException(400, "GoData no está configurado")
    client = GoDataClient(base_url=url, username=user, password=pwd, outbreak_id=outbreak_id)
    result = client.test_connection()
    # Normalize response to include success/message for frontend compatibility
    if result.get("status") == "ok":
        result["success"] = True
        result["message"] = f"Conexión exitosa. Autenticado correctamente."
    else:
        result["success"] = False
        result["message"] = result.get("detail", "Error de conexión")
    return result


@app.get("/api/godata/outbreaks")
def godata_list_outbreaks(x_api_key: str = Header(None)):
    """Lista brotes disponibles en GoData."""
    verify_api_key(x_api_key)
    url, user, pwd, _ = get_godata_credentials()
    if not url:
        raise HTTPException(400, "GoData no está configurado")
    client = GoDataClient(base_url=url, username=user, password=pwd)
    try:
        return client.get_outbreaks()
    except Exception as e:
        raise HTTPException(500, f"Error obteniendo brotes de GoData: {e}")


@app.get("/api/godata/users")
def godata_list_users(
    limit: int = 100, offset: int = 0,
    x_api_key: str = Header(None)
):
    """Lista usuarios visibles en GoData."""
    verify_api_key(x_api_key)
    url, user, pwd, _ = get_godata_credentials()
    if not url:
        raise HTTPException(400, "GoData no está configurado")
    client = GoDataClient(base_url=url, username=user, password=pwd)
    try:
        return client._get(client._api_url("/users"), params={"limit": limit, "skip": offset})
    except Exception as e:
        raise HTTPException(500, f"Error obteniendo usuarios de GoData: {e}")


@app.get("/api/godata/templates")
def godata_list_templates(
    limit: int = 100, offset: int = 0,
    x_api_key: str = Header(None)
):
    """Lista templates disponibles en GoData."""
    verify_api_key(x_api_key)
    url, user, pwd, outbreak_id = get_godata_credentials()
    if not url:
        raise HTTPException(400, "GoData no está configurado")
    client = GoDataClient(base_url=url, username=user, password=pwd, outbreak_id=outbreak_id)
    try:
        return client._get(client._api_url("/templates"), params={"limit": limit, "skip": offset})
    except Exception as e:
        raise HTTPException(500, f"Error obteniendo templates de GoData: {e}")


@app.get("/api/godata/reference-data")
def godata_reference_data(
    category: str = None,
    x_api_key: str = Header(None)
):
    """Lista datos de referencia de GoData."""
    verify_api_key(x_api_key)
    url, user, pwd, _ = get_godata_credentials()
    if not url:
        raise HTTPException(400, "GoData no está configurado")
    client = GoDataClient(base_url=url, username=user, password=pwd)
    try:
        return client.get_reference_data(category=category)
    except Exception as e:
        raise HTTPException(500, f"Error obteniendo reference data de GoData: {e}")


@app.get("/api/godata/queue")
def godata_queue_endpoint(
    estado: str = None, limit: int = 100,
    x_api_key: str = Header(None)
):
    """Cola de sincronización GoData."""
    verify_api_key(x_api_key)
    return godata_get_queue(estado=estado, limit=limit)


@app.post("/api/godata/queue/enqueue-all")
def godata_enqueue_all(x_api_key: str = Header(None)):
    """Encolar todos los registros pendientes para GoData."""
    verify_api_key(x_api_key)
    return godata_enqueue_pending()


@app.post("/api/godata/queue/approve")
async def godata_approve(request: Request, x_api_key: str = Header(None)):
    """Aprobar registros para sincronización con GoData."""
    verify_api_key(x_api_key)
    body = await request.json()
    ids = body.get("ids", [])
    if not ids:
        raise HTTPException(400, "Debe enviar lista de ids")
    return godata_approve_records(ids)


@app.post("/api/godata/sync/{registro_id}")
def godata_sync_single(registro_id: str, x_api_key: str = Header(None)):
    """Sincronizar un registro individual con GoData."""
    verify_api_key(x_api_key)
    from database import get_registro_by_id

    # Obtener registro
    record = get_registro_by_id(registro_id)
    if not record:
        raise HTTPException(404, "Registro no encontrado")

    # Claim
    if not try_claim_for_sync(registro_id):
        raise HTTPException(409, "Registro no está aprobado o ya se está sincronizando")

    # Mapear y enviar
    url, user, pwd, outbreak_id = get_godata_credentials()
    if not url:
        godata_mark_error(registro_id, "GoData no está configurado")
        raise HTTPException(400, "GoData no está configurado")

    client = GoDataClient(base_url=url, username=user, password=pwd, outbreak_id=outbreak_id)
    try:
        # If this record was previously synced (re-sync after error), skip creation
        record_godata_id = record.get("godata_case_id", "")
        if record_godata_id and not record_godata_id.startswith("DRYRUN"):
            mark_synced(registro_id, record_godata_id)
            return {"status": "duplicate", "godata_case_id": record_godata_id}

        # Verificar duplicado por registro_id
        existing = client.find_case_by_visual_id(registro_id)
        if existing:
            godata_mark_duplicate(registro_id, existing.get("id", ""))
            return {"status": "duplicate", "godata_case_id": existing.get("id")}

        # Do NOT set visualId — let GoData auto-assign from its SR-9999 mask
        case_payload = map_record_to_godata(record)

        # Validate payload before sending
        warnings = validate_godata_payload(case_payload)
        if warnings:
            logger.warning(f"GoData payload warnings for {registro_id}: {warnings}")

        result = client.create_case(case_payload)
        godata_case_id = result.get("id", "")

        # Read the visualId that GoData auto-assigned
        visual_id = result.get("visualId", "")
        if visual_id:
            save_visual_id(registro_id, visual_id)

        # Agregar resultados de laboratorio
        lab_results = map_lab_samples_to_godata(record)
        lab_count = 0
        for lab in lab_results:
            try:
                client.add_lab_result(godata_case_id, lab)
                lab_count += 1
            except Exception as e:
                logger.warning(f"GoData lab result error for {registro_id}: {e}")

        mark_synced(registro_id, godata_case_id)
        return {
            "status": "synced",
            "godata_case_id": godata_case_id,
            "visual_id": visual_id,
            "lab_results_sent": lab_count,
            "dry_run": result.get("dry_run", False),
        }
    except Exception as e:
        godata_mark_error(registro_id, str(e))
        raise HTTPException(500, f"Error sincronizando con GoData: {e}")


@app.post("/api/godata/sync-batch")
def godata_sync_batch(x_api_key: str = Header(None)):
    """Sincronizar todos los registros aprobados con GoData."""
    verify_api_key(x_api_key)
    from database import get_registro_by_id

    # Recover any stuck syncs before starting batch
    recover_stuck_syncs()

    ids = get_approved_for_sync(limit=100)
    if not ids:
        return {"processed": 0, "message": "No hay registros aprobados"}

    url, user, pwd, outbreak_id = get_godata_credentials()
    if not url:
        raise HTTPException(400, "GoData no está configurado")

    client = GoDataClient(base_url=url, username=user, password=pwd, outbreak_id=outbreak_id)
    results = {"processed": 0, "success": 0, "errors": 0, "duplicates": 0}

    for registro_id in ids:
        results["processed"] += 1
        if not try_claim_for_sync(registro_id):
            continue

        record = get_registro_by_id(registro_id)
        if not record:
            godata_mark_error(registro_id, "Registro no encontrado en BD")
            results["errors"] += 1
            continue

        try:
            # If this record was previously synced (re-sync after error), skip creation
            record_godata_id = record.get("godata_case_id", "")
            if record_godata_id and not record_godata_id.startswith("DRYRUN"):
                mark_synced(registro_id, record_godata_id)
                results["duplicates"] += 1
                continue

            # Verificar duplicado por registro_id
            existing = client.find_case_by_visual_id(registro_id)
            if existing:
                godata_mark_duplicate(registro_id, existing.get("id", ""))
                results["duplicates"] += 1
                continue

            # Do NOT set visualId — let GoData auto-assign from its SR-9999 mask
            case_payload = map_record_to_godata(record)

            # Validate payload before sending
            warnings = validate_godata_payload(case_payload)
            if warnings:
                logger.warning(f"GoData payload warnings for {registro_id}: {warnings}")

            result = client.create_case(case_payload)
            godata_case_id = result.get("id", "")

            # Read the visualId that GoData auto-assigned
            visual_id = result.get("visualId", "")
            if visual_id:
                save_visual_id(registro_id, visual_id)

            lab_results = map_lab_samples_to_godata(record)
            for lab in lab_results:
                try:
                    client.add_lab_result(godata_case_id, lab)
                except Exception as e:
                    logger.warning(f"GoData lab error {registro_id}: {e}")

            mark_synced(registro_id, godata_case_id)
            results["success"] += 1
        except Exception as e:
            godata_mark_error(registro_id, str(e))
            results["errors"] += 1
            logger.error(f"GoData sync error for {registro_id}: {e}")

    return results


@app.get("/api/godata/status/{registro_id}")
def godata_status(registro_id: str, x_api_key: str = Header(None)):
    """Estado de sincronización GoData de un registro."""
    verify_api_key(x_api_key)
    return godata_get_sync_status(registro_id)


@app.get("/api/godata/preview/{registro_id}")
def godata_preview(registro_id: str, x_api_key: str = Header(None)):
    """Preview del payload GoData que se enviaría (sin enviar)."""
    verify_api_key(x_api_key)
    from database import get_registro_by_id
    record = get_registro_by_id(registro_id)
    if not record:
        raise HTTPException(404, "Registro no encontrado")
    case_payload = map_record_to_godata(record)
    lab_payloads = map_lab_samples_to_godata(record)
    return {
        "case": case_payload,
        "lab_results": lab_payloads,
        "lab_results_count": len(lab_payloads),
    }


# ═══════════════════════════════════════════════════════════
# GoData 2-Phase Sync Endpoints
# ═══════════════════════════════════════════════════════════

@app.post("/api/godata/send-fase1/{registro_id}")
def godata_send_fase1(registro_id: str, x_api_key: str = Header(None)):
    """Phase 1: Create case in GoData with basic data (patient, symptoms, vaccination).
    Classification forced to SUSPECT. Lab results NOT included."""
    verify_api_key(x_api_key)
    from database import get_registro_by_id

    record = get_registro_by_id(registro_id)
    if not record:
        raise HTTPException(404, "Registro no encontrado")

    # Claim for sync (must be in 'aprobado' or 'error_fase1' state)
    if not try_claim_for_sync(registro_id):
        # Also allow error_fase1 retries
        from godata_queue import get_sync_status
        status = get_sync_status(registro_id)
        if status.get("estado") == "error_fase1":
            # Manually set to sincronizando for retry
            import sqlite3
            from config import DB_PATH
            conn = sqlite3.connect(DB_PATH, timeout=30)
            try:
                conn.execute("""
                    UPDATE godata_queue
                    SET estado = 'sincronizando', intentos = intentos + 1, updated_at = datetime('now')
                    WHERE registro_id = ? AND estado = 'error_fase1'
                """, (registro_id,))
                conn.commit()
            finally:
                conn.close()
        else:
            raise HTTPException(409, f"Registro no está aprobado para fase 1 (estado: {status.get('estado')})")

    url, user, pwd, outbreak_id = get_godata_credentials()
    if not url:
        mark_error_fase1(registro_id, "GoData no está configurado")
        raise HTTPException(400, "GoData no está configurado")

    client = GoDataClient(base_url=url, username=user, password=pwd, outbreak_id=outbreak_id)
    try:
        # Check if already has a GoData case ID (skip creation)
        record_godata_id = record.get("godata_case_id", "")
        if record_godata_id and not record_godata_id.startswith("DRYRUN"):
            mark_fase1_sent(registro_id, record_godata_id)
            return {"status": "already_exists", "godata_case_id": record_godata_id}

        # Check for duplicates
        existing = client.find_case_by_visual_id(registro_id)
        if existing:
            godata_mark_duplicate(registro_id, existing.get("id", ""))
            return {"status": "duplicate", "godata_case_id": existing.get("id")}

        # Build Phase 1 payload — do NOT set visualId, let GoData auto-assign
        case_payload = map_record_to_godata_fase1(record)

        warnings = validate_godata_payload(case_payload)
        if warnings:
            logger.warning(f"GoData fase1 payload warnings for {registro_id}: {warnings}")

        result = client.create_case(case_payload)
        godata_case_id = result.get("id", "")

        # Read the visualId that GoData auto-assigned
        visual_id = result.get("visualId", "")
        if visual_id:
            save_visual_id(registro_id, visual_id)
        mark_fase1_sent(registro_id, godata_case_id)

        return {
            "status": "fase1_sent",
            "godata_case_id": godata_case_id,
            "visual_id": visual_id,
            "warnings": warnings,
            "dry_run": result.get("dry_run", False),
        }
    except Exception as e:
        mark_error_fase1(registro_id, str(e))
        raise HTTPException(500, f"Error en fase 1 GoData: {e}")


@app.post("/api/godata/send-fase1-batch")
def godata_send_fase1_batch(x_api_key: str = Header(None)):
    """Phase 1 batch: Create all approved cases in GoData with basic data."""
    verify_api_key(x_api_key)
    from database import get_registro_by_id

    recover_stuck_syncs()

    ids = get_approved_for_sync(limit=100)
    if not ids:
        return {"processed": 0, "message": "No hay registros aprobados para fase 1"}

    url, user, pwd, outbreak_id = get_godata_credentials()
    if not url:
        raise HTTPException(400, "GoData no está configurado")

    client = GoDataClient(base_url=url, username=user, password=pwd, outbreak_id=outbreak_id)
    results = {"processed": 0, "success": 0, "errors": 0, "duplicates": 0, "details": []}

    for registro_id in ids:
        results["processed"] += 1
        if not try_claim_for_sync(registro_id):
            continue

        record = get_registro_by_id(registro_id)
        if not record:
            mark_error_fase1(registro_id, "Registro no encontrado en BD")
            results["errors"] += 1
            continue

        try:
            record_godata_id = record.get("godata_case_id", "")
            if record_godata_id and not record_godata_id.startswith("DRYRUN"):
                mark_fase1_sent(registro_id, record_godata_id)
                results["duplicates"] += 1
                continue

            existing = client.find_case_by_visual_id(registro_id)
            if existing:
                godata_mark_duplicate(registro_id, existing.get("id", ""))
                results["duplicates"] += 1
                continue

            # Do NOT set visualId — let GoData auto-assign
            case_payload = map_record_to_godata_fase1(record)

            warnings = validate_godata_payload(case_payload)
            if warnings:
                logger.warning(f"GoData fase1 warnings for {registro_id}: {warnings}")

            result = client.create_case(case_payload)
            godata_case_id = result.get("id", "")

            # Read the visualId that GoData auto-assigned
            visual_id = result.get("visualId", "")
            if visual_id:
                save_visual_id(registro_id, visual_id)
            mark_fase1_sent(registro_id, godata_case_id)
            results["success"] += 1
            results["details"].append({
                "registro_id": registro_id,
                "godata_case_id": godata_case_id,
                "visual_id": visual_id,
                "dry_run": result.get("dry_run", False),
            })
        except Exception as e:
            mark_error_fase1(registro_id, str(e))
            results["errors"] += 1
            logger.error(f"GoData fase1 error for {registro_id}: {e}")

    return results


@app.post("/api/godata/send-fase2/{registro_id}")
def godata_send_fase2(registro_id: str, x_api_key: str = Header(None)):
    """Phase 2: Update case in GoData with lab results, classification, and condition.
    Sends ALL fields (Phase 1 + Phase 2) because PUT replaces questionnaireAnswers."""
    verify_api_key(x_api_key)
    from database import get_registro_by_id

    record = get_registro_by_id(registro_id)
    if not record:
        raise HTTPException(404, "Registro no encontrado")

    # Claim for Phase 2
    if not try_claim_for_fase2(registro_id):
        from godata_queue import get_sync_status
        status = get_sync_status(registro_id)
        raise HTTPException(409, f"Registro no está listo para fase 2 (estado: {status.get('estado')})")

    # Get the godata_case_id assigned during Phase 1
    from godata_queue import get_sync_status
    queue_status = get_sync_status(registro_id)
    godata_case_id = queue_status.get("godata_case_id", "")
    if not godata_case_id:
        mark_error_fase2(registro_id, "No hay godata_case_id — fase 1 no se completó")
        raise HTTPException(400, "Registro no tiene godata_case_id. Ejecute fase 1 primero.")

    url, user, pwd, outbreak_id = get_godata_credentials()
    if not url:
        mark_error_fase2(registro_id, "GoData no está configurado")
        raise HTTPException(400, "GoData no está configurado")

    client = GoDataClient(base_url=url, username=user, password=pwd, outbreak_id=outbreak_id)
    try:
        # Build Phase 2 payload (COMPLETE data including classification + lab)
        case_payload = map_record_to_godata_fase2(record)

        # Keep the visualId from Phase 1
        visual_id = queue_status.get("godata_visual_id", "")
        if visual_id:
            case_payload["visualId"] = visual_id

        warnings = validate_godata_payload(case_payload)
        if warnings:
            logger.warning(f"GoData fase2 payload warnings for {registro_id}: {warnings}")

        # PUT update replaces the entire case
        result = client.update_case(godata_case_id, case_payload)

        mark_complete(registro_id)

        return {
            "status": "complete",
            "godata_case_id": godata_case_id,
            "warnings": warnings,
            "dry_run": result.get("dry_run", False),
        }
    except Exception as e:
        mark_error_fase2(registro_id, str(e))
        raise HTTPException(500, f"Error en fase 2 GoData: {e}")


@app.post("/api/godata/send-fase2-batch")
def godata_send_fase2_batch(x_api_key: str = Header(None)):
    """Phase 2 batch: Update all subido_fase1 cases with lab results and classification."""
    verify_api_key(x_api_key)
    from database import get_registro_by_id

    recover_stuck_syncs()

    pending = get_fase1_pending(limit=100)
    if not pending:
        return {"processed": 0, "message": "No hay registros en subido_fase1 para fase 2"}

    url, user, pwd, outbreak_id = get_godata_credentials()
    if not url:
        raise HTTPException(400, "GoData no está configurado")

    client = GoDataClient(base_url=url, username=user, password=pwd, outbreak_id=outbreak_id)
    results = {"processed": 0, "success": 0, "errors": 0, "details": []}

    for item in pending:
        registro_id = item["registro_id"]
        godata_case_id = item["godata_case_id"]
        results["processed"] += 1

        if not godata_case_id:
            mark_error_fase2(registro_id, "No hay godata_case_id")
            results["errors"] += 1
            continue

        if not try_claim_for_fase2(registro_id):
            continue

        record = get_registro_by_id(registro_id)
        if not record:
            mark_error_fase2(registro_id, "Registro no encontrado en BD")
            results["errors"] += 1
            continue

        try:
            case_payload = map_record_to_godata_fase2(record)

            # Keep the visualId from Phase 1
            from godata_queue import get_sync_status
            queue_status = get_sync_status(registro_id)
            visual_id = queue_status.get("godata_visual_id", "")
            if visual_id:
                case_payload["visualId"] = visual_id

            warnings = validate_godata_payload(case_payload)
            if warnings:
                logger.warning(f"GoData fase2 warnings for {registro_id}: {warnings}")

            result = client.update_case(godata_case_id, case_payload)

            mark_complete(registro_id)
            results["success"] += 1
            results["details"].append({
                "registro_id": registro_id,
                "godata_case_id": godata_case_id,
                "dry_run": result.get("dry_run", False),
            })
        except Exception as e:
            mark_error_fase2(registro_id, str(e))
            results["errors"] += 1
            logger.error(f"GoData fase2 error for {registro_id}: {e}")

    return results


@app.post("/api/godata/requeue-update/{registro_id}")
def godata_requeue_update(registro_id: str, x_api_key: str = Header(None)):
    """Move completed case back to subido_fase1 for re-update."""
    verify_api_key(x_api_key)
    if requeue_for_update(registro_id):
        return {"status": "ok", "message": "Caso listo para re-actualización"}
    raise HTTPException(400, "Caso no está en estado 'completo' o no existe")


@app.post("/api/godata/queue/unapprove")
async def godata_unapprove(request: Request, x_api_key: str = Header(None)):
    """Move approved records back to pending."""
    verify_api_key(x_api_key)
    body = await request.json()
    ids = body.get("ids", [])
    if not ids:
        raise HTTPException(400, "Debe enviar lista de ids")
    return unapprove_records(ids)


@app.get("/api/godata/case/{registro_id}")
def godata_get_case(registro_id: str, x_api_key: str = Header(None)):
    """Fetch case data from GoData for comparison."""
    verify_api_key(x_api_key)
    status = godata_get_sync_status(registro_id)
    godata_case_id = status.get("godata_case_id", "")
    if not godata_case_id:
        raise HTTPException(404, "Caso no tiene GoData ID")

    url, user, pwd, outbreak_id = get_godata_credentials()
    if not url:
        raise HTTPException(400, "GoData no está configurado")

    client = GoDataClient(base_url=url, username=user, password=pwd, outbreak_id=outbreak_id)
    try:
        case = client.get_case(godata_case_id)
        return {
            "godata_case": case,
            "godata_case_id": godata_case_id,
            "visual_id": case.get("visualId"),
            "qa_keys": len(case.get("questionnaireAnswers", {})),
        }
    except Exception as e:
        raise HTTPException(500, f"Error consultando GoData: {e}")


# ─── Reportes / Pendientes ─────────────────────────────────

@app.get("/api/reportes/pendientes-resultado")
def reporte_pendientes_resultado(x_api_key: str = Header(None)):
    """Pacientes con muestra recolectada pero sin resultado de laboratorio."""
    verify_api_key(x_api_key)
    from database import get_connection
    conn = get_connection()
    try:
        rows = conn.execute("""
            SELECT registro_id, nombres, apellidos, afiliacion,
                   fecha_notificacion, unidad_medica,
                   muestra_suero, muestra_suero_fecha,
                   muestra_hisopado, muestra_hisopado_fecha,
                   muestra_orina, muestra_orina_fecha,
                   clasificacion_caso, recolecto_muestra,
                   godata_sync_status, godata_case_id,
                   resultado_prueba, resultado_igm_cualitativo,
                   resultado_igg_cualitativo, resultado_pcr_orina,
                   resultado_pcr_hisopado, lab_muestras_json
            FROM registros
            WHERE recolecto_muestra = 'SI'
              AND (clasificacion_caso IN ('SOSPECHOSO', 'PENDIENTE', '')
                   OR clasificacion_caso IS NULL)
            ORDER BY fecha_notificacion DESC
        """).fetchall()
        return {"data": [dict(r) for r in rows], "total": len(rows)}
    finally:
        conn.close()



# ─── DDRISS Reporting ───────────────────────────────────────

@app.get("/api/reportes/ddriss-list")
def reporte_ddriss_list(x_api_key: str = Header(None)):
    """Get the list of all DDRISS names."""
    verify_api_key(x_api_key)
    from ddriss_mapping import DDRISS_LIST
    return {"ddriss": DDRISS_LIST}


@app.get("/api/reportes/ddriss-counts")
def reporte_ddriss_counts(
    fecha_inicio: str = None,
    fecha_fin: str = None,
    x_api_key: str = Header(None),
):
    """Get count of records per DDRISS, optionally filtered by date range."""
    verify_api_key(x_api_key)
    from ddriss_mapping import get_ddriss, DDRISS_LIST, _parse_date_to_iso
    from database import get_connection

    conn = get_connection()
    try:
        rows = conn.execute(
            "SELECT departamento_residencia, municipio_residencia, fecha_notificacion FROM registros"
        ).fetchall()

        # Parse date filters
        iso_inicio = fecha_inicio or ""
        iso_fin = fecha_fin or ""

        counts = {d: 0 for d in DDRISS_LIST}
        total = 0
        for row in rows:
            # Date filter (handles DD/MM/YYYY and YYYY-MM-DD)
            if iso_inicio or iso_fin:
                fecha_iso = _parse_date_to_iso(row['fecha_notificacion'] or '')
                if not fecha_iso:
                    continue
                if iso_inicio and fecha_iso < iso_inicio:
                    continue
                if iso_fin and fecha_iso > iso_fin:
                    continue

            total += 1
            ddriss = get_ddriss(row['departamento_residencia'] or '', row['municipio_residencia'] or '')
            if ddriss in counts:
                counts[ddriss] += 1
            else:
                counts[ddriss] = counts.get(ddriss, 0) + 1

        # Sort by count descending, then name
        sorted_counts = dict(sorted(counts.items(), key=lambda x: (-x[1], x[0])))
        return {"counts": sorted_counts, "total": total}
    finally:
        conn.close()


@app.post("/api/reportes/fichas-por-ddriss")
async def reporte_fichas_por_ddriss(request: Request, x_api_key: str = Header(None)):
    """Generate bulk fichas filtered by DDRISS and date range.
    Body: {"ddriss": "GUATEMALA SUR", "fecha_inicio": "2026-01-01", "fecha_fin": "2026-03-31", "format": "merged"|"zip"}
    """
    verify_api_key(x_api_key)
    body = await request.json()
    ddriss = body.get("ddriss", "").strip()
    fecha_inicio = body.get("fecha_inicio", "")
    fecha_fin = body.get("fecha_fin", "")
    fmt = body.get("format", "merged")

    if not ddriss:
        raise HTTPException(400, "Debe seleccionar un DDRISS")

    from ddriss_mapping import get_ddriss, _parse_date_to_iso
    from database import get_connection

    conn = get_connection()
    try:
        rows = conn.execute("SELECT * FROM registros").fetchall()

        # Filter by date and DDRISS
        records = []
        for row in rows:
            r = dict(row)
            # Date filter
            if fecha_inicio or fecha_fin:
                fecha_iso = _parse_date_to_iso(r.get('fecha_notificacion', '') or '')
                if not fecha_iso:
                    continue
                if fecha_inicio and fecha_iso < fecha_inicio:
                    continue
                if fecha_fin and fecha_iso > fecha_fin:
                    continue

            # "TODAS" means no DDRISS filter
            if ddriss.upper() == 'TODAS':
                records.append(r)
            else:
                record_ddriss = get_ddriss(
                    r.get('departamento_residencia', ''),
                    r.get('municipio_residencia', '')
                )
                if record_ddriss == ddriss:
                    records.append(r)

        if not records:
            raise HTTPException(404, f"No se encontraron registros para {ddriss} en el rango de fechas indicado")

        from pdf_ficha_v2 import generar_fichas_v2_bulk
        merge = (fmt == "merged")
        result_bytes = generar_fichas_v2_bulk(records, merge=merge)

        ext = "pdf" if merge else "zip"
        mime = "application/pdf" if merge else "application/zip"
        safe_name = ddriss.replace(' ', '_')
        ts = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"fichas_{safe_name}_{ts}.{ext}"

        return StreamingResponse(
            io.BytesIO(result_bytes),
            media_type=mime,
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )
    finally:
        conn.close()



if __name__ == "__main__":
    import uvicorn
    uvicorn.run("main:app", host="0.0.0.0", port=PORT, reload=True)

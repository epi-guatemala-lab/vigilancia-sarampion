"""
PDF Generator for MSPAS Sarampion/Rubeola Form — Excel Template Version.

Opens the official MSPAS 2026 Excel template (.xlsx), fills in patient data,
and converts to PDF via LibreOffice.  Produces a pixel-perfect 2-page replica
because the template already contains all borders, fonts, colors, and merged cells.

Template: ficha_sarampion_template.xlsx — Sheet "Hoja1" — 108 rows x 20 cols (A-T)
          384 merged cells, page break after row 57 (end of Section 4).

Public API:
    generar_ficha_v2_pdf(data: dict) -> bytes
    generar_fichas_v2_bulk(records: list, merge: bool = True) -> bytes
"""
import copy
import io
import json
import logging
import os
import shutil
import subprocess
import tempfile
import zipfile
from datetime import datetime

import openpyxl
from openpyxl.drawing.image import Image as XlImage
from openpyxl.worksheet.pagebreak import Break, RowBreak

logger = logging.getLogger(__name__)

ASSETS_DIR = os.path.join(os.path.dirname(__file__), "assets")
TEMPLATE_PATH = os.path.join(ASSETS_DIR, "ficha_sarampion_template.xlsx")
LOGO_PATH = os.path.join(ASSETS_DIR, "sello_mspas.png")

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _g(data: dict, key: str, default: str = "") -> str:
    """Safe get from data dict; treats None/NaN/NULL as empty."""
    val = data.get(key)
    if val is None:
        return default
    s = str(val).strip()
    if s.upper() in ("NONE", "NULL", "NAN", ""):
        return default
    return s


def _parse_date(date_str):
    """Parse a date string and return (DD, MM, YYYY) as strings."""
    if not date_str:
        return ("", "", "")
    date_str = str(date_str).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%Y-%m-%dT%H:%M:%S", "%d-%m-%Y"):
        try:
            dt = datetime.strptime(date_str[:10], fmt)
            return (f"{dt.day:02d}", f"{dt.month:02d}", str(dt.year))
        except (ValueError, TypeError):
            continue
    return ("", "", "")


def _chk(value) -> bool:
    """Check if a value is truthy / 'SI'."""
    if value is None:
        return False
    return str(value).strip().upper() in ("SI", "SÍ", "S", "1", "TRUE", "X", "YES")


def _is_no(value) -> bool:
    if value is None:
        return False
    return str(value).strip().upper() in ("NO", "N", "0", "FALSE")


def _is_desc(value) -> bool:
    if value is None:
        return False
    return str(value).strip().upper() in ("DESCONOCIDO", "DESC", "D", "UNKNOWN")


def _strip_accents(s: str) -> str:
    """Remove diacritics/accents from a string for robust matching."""
    import unicodedata
    return ''.join(
        c for c in unicodedata.normalize('NFD', s)
        if unicodedata.category(c) != 'Mn'
    )


def _safe_json(val):
    if not val:
        return []
    if isinstance(val, (list, dict)):
        return val
    try:
        return json.loads(val)
    except Exception:
        return []


# ---------------------------------------------------------------------------
# Lab result abbreviation (cells are narrow)
# ---------------------------------------------------------------------------

_LAB_ABBREV = {
    'REACTIVO': 'R+', 'NO REACTIVO': 'NR', 'INDETERMINADO': 'IND',
    'POSITIVO': 'POS', 'NEGATIVO': 'NEG', 'PENDIENTE': 'PEND',
    'NO APLICA': 'N/A', 'N/A': 'N/A',
}


def _lab_val(val: str) -> str:
    """Abbreviate lab result for narrow cells."""
    if not val:
        return val
    return _LAB_ABBREV.get(val.upper().strip(), val)


def _trunc(text: str, max_len: int) -> str:
    """Truncate text to max_len chars, adding ... if truncated."""
    if not text or len(str(text)) <= max_len:
        return text
    return str(text)[:max_len-2] + '..'

# ---------------------------------------------------------------------------
# Cell writing helpers
# ---------------------------------------------------------------------------

def _merged_width(ws, row, col):
    """Calculate approximate character width of a cell, accounting for merges."""
    for mc in ws.merged_cells.ranges:
        if ws.cell(row=row, column=col).coordinate == str(mc).split(':')[0]:
            total = 0
            for c in range(mc.min_col, mc.max_col + 1):
                from openpyxl.utils import get_column_letter
                letter = get_column_letter(c)
                total += ws.column_dimensions[letter].width or 8
            return total
    from openpyxl.utils import get_column_letter
    letter = get_column_letter(col)
    return ws.column_dimensions[letter].width or 8


def _write(ws, row, col, value):
    """Write a value to a cell, preserving existing formatting.
    If the cell is part of a merged range (not top-left), skip silently.

    Text handling strategy (hybrid):
      - Short text (fits cell width): normal display
      - Medium text: wrap_text=True (readable, may increase row height)
      - Very long text (>2x cell width): shrink_to_fit as fallback
    """
    from openpyxl.cell.cell import MergedCell
    from openpyxl.styles import Font, Alignment
    cell = ws.cell(row=row, column=col)
    if isinstance(cell, MergedCell):
        logger.warning("Attempted write to merged cell R%d:C%d, skipping", row, col)
        return
    cell.value = value
    if value and str(value).strip():
        text_len = len(str(value))
        cell_width = _merged_width(ws, row, col)

        old_font = cell.font
        # Use slightly smaller font for data to distinguish from labels
        cell.font = Font(
            name=old_font.name or 'Calibri',
            size=old_font.size or 9,
            bold=True,
        )
        old_align = cell.alignment

        # Approximate: 1 character ~ 1 width unit for Calibri 9pt
        fits = text_len <= cell_width * 0.9
        very_long = text_len > cell_width * 2.0

        if fits:
            # Short text — fits naturally
            cell.alignment = Alignment(
                horizontal=old_align.horizontal or 'left',
                vertical=old_align.vertical or 'center',
            )
        elif very_long:
            # Very long text — shrink to fit as last resort
            cell.alignment = Alignment(
                horizontal=old_align.horizontal or 'left',
                vertical=old_align.vertical or 'center',
                shrink_to_fit=True,
            )
        else:
            # Medium text — wrap to stay readable
            cell.alignment = Alignment(
                horizontal=old_align.horizontal or 'left',
                vertical=old_align.vertical or 'center',
                wrap_text=True,
            )


def _check(ws, row, col, condition):
    """Mark checkbox: prepend checkmark to existing label text, or write checkmark if cell is empty.
    Makes the checkbox BOLD for visibility. Uses wrap_text for long labels."""
    if condition:
        from openpyxl.cell.cell import MergedCell
        from openpyxl.styles import Font, Alignment
        cell = ws.cell(row=row, column=col)
        if isinstance(cell, MergedCell):
            logger.warning("Attempted check on merged cell R%d:C%d, skipping", row, col)
            return
        existing = cell.value
        if existing and str(existing).strip():
            text = str(existing).strip()
            if not text.startswith('\u2612'):
                cell.value = f"\u2612 {text}"
        else:
            cell.value = "\u2612"
        # Make the checkbox text BOLD for visibility
        old_font = cell.font
        cell.font = Font(
            name=old_font.name or 'Calibri',
            size=old_font.size or 9,
            bold=True,
            color='000000',
        )
        # Ensure long checkbox labels wrap instead of overflowing
        final_text = str(cell.value)
        cell_w = _merged_width(ws, row, col)
        if len(final_text) > cell_w * 0.9:
            old_align = cell.alignment
            cell.alignment = Alignment(
                horizontal=old_align.horizontal or 'center',
                vertical=old_align.vertical or 'center',
                wrap_text=True,
            )


def _write_date(ws, row, col_d, col_m, col_y, date_str):
    """Write DD, MM, YYYY into three separate cells for a date field.
    Centers the values since date cells are narrow."""
    from openpyxl.cell.cell import MergedCell
    from openpyxl.styles import Font, Alignment
    dd, mm, yyyy = _parse_date(date_str)
    if dd:
        for col, val in [(col_d, dd), (col_m, mm), (col_y, yyyy)]:
            cell = ws.cell(row=row, column=col)
            if isinstance(cell, MergedCell):
                continue
            cell.value = val
            cell.font = Font(name='Calibri', size=9, bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')


def _check_si_no_desc(ws, row, col_si, col_no, col_desc, value):
    """Check the appropriate SI/NO/Desc checkbox based on value."""
    _check(ws, row, col_si, _chk(value))
    _check(ws, row, col_no, _is_no(value))
    _check(ws, row, col_desc, _is_desc(value))


def _check_si_no(ws, row, col_si, col_no, value):
    """Check SI or NO checkbox."""
    _check(ws, row, col_si, _chk(value))
    _check(ws, row, col_no, _is_no(value))


# ---------------------------------------------------------------------------
# Symptom helper (Si/No/Desc)
# ---------------------------------------------------------------------------

def _write_symptom(ws, row, col_si, col_no, col_desc, value):
    """Write a symptom checkbox row: SI / NO / Desc."""
    if value is None or str(value).strip() == "":
        return
    v = str(value).strip().upper()
    if v in ("SI", "SÍ", "S", "1", "TRUE", "X", "YES"):
        _write(ws, row, col_si, "\u2612")
    elif v in ("NO", "N", "0", "FALSE"):
        _write(ws, row, col_no, "\u2612")
    elif v in ("DESCONOCIDO", "DESC", "D", "UNKNOWN"):
        _write(ws, row, col_desc, "\u2612")


# ---------------------------------------------------------------------------
# Fill template with data — NEW 2026 TEMPLATE (108 rows, Sheet "Hoja1")
# ---------------------------------------------------------------------------

def _fill_template(ws, d: dict):
    """Fill the 'Hoja1' worksheet with patient data.

    Cell positions are based on the 2026 template (108 rows x 19+ cols A-S).

    Template layout (row ranges):
      R1-7:   Header + instructions
      R8-9:   Diagnostico de Sospecha
      R10:    Section 1 header
      R11-19: Unidad Notificadora
      R21:    Section 2 header
      R22-33: Informacion del Paciente
      R35:    Section 3 header
      R36-44: Antecedentes Medicos y Vacunacion
      R46:    Section 4 header
      R47-57: Datos Clinicos
      R61:    Section 5 header
      R62-69: Factores de Riesgo
      R71:    Section 6 header
      R72-76: Acciones de Respuesta
      R77:    Section 7 header
      R78-92: Laboratorio
      R93:    Classification header
      R94-108: Clasificacion Final
    """

    from openpyxl.cell.cell import MergedCell
    from openpyxl.styles import Font, Alignment

    # ===== ROWS 8-9: DIAGNOSTICO DE SOSPECHA =====
    # R8: A8:D9='Diagnostico de Sospecha' (merged label)
    #     E8:F8='Sarampion', G8:H8='Rubeola', I8:J8='Dengue',
    #     K8:M8='Otra Arbovirosis', N8(not merged)='Especifique', P8:S8=data
    # R9: E9(not merged)='Otra febril exantematica',
    #     I9:J9='Especifique', K9:M9=data,
    #     N9:S9='Caso altamente sospechoso de Sarampion'
    diag = _strip_accents(_g(d, 'diagnostico_sospecha', _g(d, 'diagnostico_registrado', '')).upper())
    is_sar = 'SARAMP' in diag or 'B05' in diag
    is_rub = 'RUBEO' in diag or 'RUBE' in diag or 'B06' in diag
    is_dengue = 'DENGUE' in diag or 'A90' in diag or 'A91' in diag
    is_arbo = 'ARBO' in diag or 'ZIKA' in diag or 'CHIK' in diag
    is_otra_febril = not (is_sar or is_rub or is_dengue or is_arbo) and bool(diag)

    _check(ws, 8, 5, is_sar)        # E8 = Sarampion
    _check(ws, 8, 7, is_rub)        # G8 = Rubeola
    _check(ws, 8, 9, is_dengue)     # I8 = Dengue
    _check(ws, 8, 11, is_arbo)      # K8 = Otra Arbovirosis

    # Especifique for arbovirosis (P8:S8 = data area)
    diag_otro = _g(d, 'diagnostico_sospecha_otro', _g(d, 'diagnostico_otro', ''))
    if is_arbo and diag_otro:
        _write(ws, 8, 16, diag_otro)
    elif is_arbo:
        arbo_text = ''
        if 'ZIKA' in diag:
            arbo_text = 'Zika'
        elif 'CHIK' in diag:
            arbo_text = 'Chikungunya'
        if arbo_text:
            _write(ws, 8, 16, arbo_text)

    # R9: Otra febril exantematica
    _check(ws, 9, 5, is_otra_febril)  # E9 = Otra febril
    if is_otra_febril and diag_otro:
        _write(ws, 9, 11, diag_otro)   # K9:M9 = Especifique data

    # Caso altamente sospechoso de Sarampion
    caso_alta = _g(d, 'caso_altamente_sospechoso', '')
    _check(ws, 9, 14, _chk(caso_alta))  # N9:S9

    # ===== SECCION 1: DATOS DE LA UNIDAD NOTIFICADORA (Rows 10-19) =====

    # R11: A11:B12='Fecha de Notificacion' (merged), C11='Dia', D11='Mes', E11='Ano'
    #      F11:J12='Direccion de Area de Salud' (merged label+data)
    #      K11:N12='Distrito de Salud' (merged label+data)
    #      O11:S12='Servicio de Salud' (merged label+data)
    # Data goes in the merged cells that also contain labels — we need separate rows.
    # Actually F11:J12 is the LABEL. Data should go below? No — it's a single merged area.
    # Looking at the template structure: the labels ARE in the merged cells.
    # For Fecha, the Dia/Mes/Ano labels are on R11, data overwrites them.
    _write_date(ws, 11, 3, 4, 5, _g(d, 'fecha_notificacion'))

    area_salud = _g(d, 'area_salud_mspas', _g(d, 'departamento_residencia', ''))
    distrito = _g(d, 'distrito_salud_mspas', _g(d, 'municipio_residencia', ''))
    servicio = _g(d, 'servicio_salud_mspas', _g(d, 'unidad_medica', ''))

    # F11:J12, K11:N12, O11:S12 are merged cells that contain labels.
    # We overwrite with "label\nDATA" and use wrap_text + top-align so
    # the label stays on line 1 and the data appears on line 2.
    from openpyxl.styles import Font, Alignment

    def _write_label_data(ws, row, col, label, data, font_size=8):
        """Write label + data into a merged label cell with proper formatting.
        Uses wrap_text for moderate data, shrink_to_fit for very long data."""
        from openpyxl.cell.cell import MergedCell
        cell = ws.cell(row=row, column=col)
        if isinstance(cell, MergedCell):
            return
        if data:
            cell.value = f"{label}\n{data}"
        else:
            cell.value = label  # keep original label
        # Calculate if data fits: cell width * 2 lines available for data
        cell_w = _merged_width(ws, row, col)
        data_len = len(data) if data else 0
        # At font 7, ~1 char per width unit. 2 rows available for data.
        fits_wrap = data_len <= cell_w * 1.8
        if fits_wrap:
            cell.font = Font(name='Calibri', size=font_size, bold=True)
            cell.alignment = Alignment(
                horizontal='center', vertical='top', wrap_text=True,
            )
        else:
            # Very long data — use smaller font + shrink
            cell.font = Font(name='Calibri', size=6, bold=True)
            cell.alignment = Alignment(
                horizontal='center', vertical='top', wrap_text=True, shrink_to_fit=True,
            )

    _write_label_data(ws, 11, 6, "Dirección de Área de Salud", area_salud, font_size=7)
    _write_label_data(ws, 11, 11, "Distrito de Salud", distrito, font_size=7)
    _write_label_data(ws, 11, 15, "Servicio de Salud", servicio, font_size=7)

    # Increase row heights for R11-R12 to accommodate 2-line content
    ws.row_dimensions[11].height = 14
    ws.row_dimensions[12].height = 14

    # R13: A13:B15='Fecha de consulta' (merged), C13='Dia', D13='Mes', E13='Ano'
    #      F13:G15='Fecha de Investigacion Domiciliaria' (merged), H13='Dia'(non-label row 14+)
    #      Wait — looking at merged ranges: C14:C15, D14:D15, E14:E15 are merged (data cells for consulta)
    #      And H14:H15, I14:I15, J14:J15 are merged (data cells for investigacion)
    # So: R13 row has labels, R14-15 merged pairs are DATA cells.
    # Fecha de consulta: data in C14, D14, E14 (merged C14:C15, D14:D15, E14:E15)
    _write_date(ws, 14, 3, 4, 5, _g(d, 'fecha_captacion', _g(d, 'fecha_registro_diagnostico', '')))

    # Fecha de Investigacion Domiciliaria: data in H14, I14, J14
    _write_date(ws, 14, 8, 9, 10, _g(d, 'fecha_visita_domiciliaria', _g(d, 'fecha_inicio_investigacion', '')))

    # Fix the label cell F13 — "Fecha de Investigación Domiciliaria" breaks badly
    # because "Investigación" (14 chars) exceeds the ~10-char cell width.
    # Reduce font to 7pt so words fit within cell width.
    label_cell = ws.cell(row=13, column=6)
    if not isinstance(label_cell, MergedCell):
        label_cell.font = Font(name='Calibri', size=7, bold=True)
        label_cell.alignment = Alignment(
            horizontal='center', vertical='center', wrap_text=True,
        )

    # K13:N13='Nombre de quien investiga' (label), O13:S13=empty data (~33 chars wide)
    # O14:S14='Cargo' — single row, no room for wrap. Use shrink for long values.
    for row_n, key in [(13, 'nom_responsable'), (14, 'cargo_responsable')]:
        val = _g(d, key)
        if val:
            cell = ws.cell(row=row_n, column=15)
            if not isinstance(cell, MergedCell):
                cell.value = val
                if len(val) <= 30:
                    cell.font = Font(name='Calibri', size=8, bold=True)
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                else:
                    cell.font = Font(name='Calibri', size=7, bold=True)
                    cell.alignment = Alignment(horizontal='left', vertical='center', shrink_to_fit=True)

    # R15: K15:M15='Telefono y Correo', N15:S15=data
    tel_correo = _g(d, 'telefono_responsable', '')
    correo = _g(d, 'correo_responsable', '')
    tel_text = ''
    if correo and tel_correo:
        tel_text = f"{tel_correo} / {correo}"
    else:
        tel_text = tel_correo or correo
    if tel_text:
        cell = ws.cell(row=15, column=14)
        if not isinstance(cell, MergedCell):
            cell.value = tel_text
            if len(tel_text) <= 35:
                cell.font = Font(name='Calibri', size=7, bold=True)
                cell.alignment = Alignment(horizontal='left', vertical='center')
            else:
                cell.font = Font(name='Calibri', size=6, bold=True)
                cell.alignment = Alignment(horizontal='left', vertical='center', shrink_to_fit=True)

    # R16: A16:C17='Seguro Social (IGSS)' (merged), D16:D17='checkbox/data'
    #      E16(not merged)='Especifique', J16:M17='Establecimiento Privado' (merged)
    #      N16:N17='checkbox', Q16:S16=data (Especifique)
    # IGSS checkbox + specify
    igss_name = _g(d, 'unidad_medica', '') if _chk(_g(d, 'es_seguro_social', 'SI')) else ''
    _check(ws, 16, 4, bool(igss_name))  # D16 = IGSS checkbox
    if igss_name:
        # E17:I17 merged = IGSS establishment name data
        # Use smaller font for long IGSS unit names to prevent overflow
        cell = ws.cell(row=17, column=5)
        if not isinstance(cell, MergedCell):
            cell.value = igss_name
            name_len = len(igss_name)
            font_sz = 7 if name_len > 35 else 8
            cell.font = Font(name='Calibri', size=font_sz, bold=True)
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        # Ensure row 17 is tall enough for wrapped IGSS name
        ws.row_dimensions[17].height = 20

    # Establecimiento Privado
    priv_name = _g(d, 'establecimiento_privado_nombre', '')
    _check(ws, 16, 14, bool(priv_name))  # N16 = Private checkbox
    if priv_name:
        _write(ws, 17, 15, priv_name)  # O17:S17

    # R18-19: Fuente de Notificacion checkboxes
    # R18: A18:C20='Fuente de Notificacion' (merged)
    #      D18:E18='Servicio de Salud', G18:H18='Laboratorio',
    #      I18:L18='Busqueda Activa Institucional', M18:N18='Busqueda Activa Comunitaria'
    #      P18:R18='Busqueda Activa Laboratorial'
    # R19: D19:E20='Investigacion de Contactos', G19:I20='Caso Reportado Comunidad'
    #      J19:L20='Auto Notificacion', M19:M20(not merged but close)='Otros'(N19)
    #      N19:O20='Otros', R19:S19='Especifique'
    fuente = _g(d, 'fuente_notificacion', '').upper()
    _check(ws, 18, 4, 'SERVICIO' in fuente or 'SALUD' in fuente)           # D18
    _check(ws, 18, 7, 'LABORATORIO' in fuente and 'ACTIVA' not in fuente)  # G18
    _check(ws, 18, 9, 'ACTIVA INSTITUCIONAL' in fuente or 'BAI' in fuente) # I18
    _check(ws, 18, 13, 'ACTIVA COMUNITARIA' in fuente or 'BAC' in fuente)  # M18
    _check(ws, 18, 16, 'ACTIVA LABORATORIAL' in fuente)                     # P18

    _check(ws, 19, 4, 'CONTACTO' in fuente or 'INVESTIGACION' in fuente)   # D19
    _check(ws, 19, 7, 'REPORTADO' in fuente or 'COMUNIDAD' in fuente)      # G19 (Caso Reportado)
    _check(ws, 19, 10, 'AUTO' in fuente or 'GRATUITO' in fuente)           # J19 (Auto Notificacion)
    _check(ws, 19, 14, 'OTRO' in fuente and 'CASO' not in fuente)          # N19 (Otros)

    # Reduce font on checked fuente notification labels (they are long)
    for r, c in [(18, 4), (18, 7), (18, 9), (18, 13), (18, 16),
                 (19, 4), (19, 7), (19, 10), (19, 14)]:
        cell = ws.cell(row=r, column=c)
        if not isinstance(cell, MergedCell) and cell.value and str(cell.value).startswith('\u2612'):
            cell.font = Font(name='Calibri', size=7, bold=True)
    fuente_otra = _g(d, 'fuente_notificacion_otra', '')
    if fuente_otra:
        _write(ws, 19, 18, fuente_otra)  # R19:S19 = Especifique data

    # P20:S20 is empty data area — could be used for extra notes
    # R20 has no labeled content

    # ===== SECCION 2: INFORMACION DEL PACIENTE (Rows 21-33) =====

    # R22: A22:B22='Nombres', C22:G22=data(nombres), H22:I22='Apellidos', J22:M22=data(apellidos)
    #      N22:N23='Sexo' (merged), O22:P23='Femenino', Q22:R23='Masculino'
    # C22:G22 (~24 chars wide) and J22:M22 (~22 chars wide) for names
    for col, key in [(3, 'nombres'), (10, 'apellidos')]:
        val = _g(d, key)
        if val:
            cell = ws.cell(row=22, column=col)
            if not isinstance(cell, MergedCell):
                cell.value = val
                font_sz = 8 if len(val) <= 22 else 7
                cell.font = Font(name='Calibri', size=font_sz, bold=True)
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

    # Ensure rows 22-23 have enough height for wrapped long names
    ws.row_dimensions[22].height = 15
    ws.row_dimensions[23].height = 12

    sexo = _g(d, 'sexo', '').upper()
    _check(ws, 22, 15, sexo in ('F', 'FEMENINO'))   # O22 = Femenino
    _check(ws, 22, 17, sexo in ('M', 'MASCULINO'))  # Q22 = Masculino

    # R24: A24:C26='Fecha de Nacimiento' (merged), D24='Dia', E24(not there?)
    # Looking at cell values: R24:C4='Dia', C5='Mes', C6='Ano'
    # But merged ranges show: A24:C26 is the label "Fecha de Nacimiento"
    # D25:D26, E25:E26, F25:F26 are merged DATA cells
    # G24:H26='Edad' (merged label), I25:I26, J25:J26, K25:K26 = DATA cells for age
    # R24:C7='Edad', C9='Dia', C10='Mes', C11='Ano' — these are age unit labels
    # L24:N25='Codigo Unico de Identificacion' (merged), O24:S26=data (CUI)
    _write_date(ws, 25, 4, 5, 6, _g(d, 'fecha_nacimiento'))  # D25, E25, F25

    # Age: values in I25, J25, K25 — but wait, R24:C9='Dia' is confusing.
    # Template shows R24:C7='Edad', R24:C9='Dia', R24:C10='Mes', R24:C11='Ano'
    # These are LABELS for age components (days/months/years as age units)
    # Data goes in I25:I26, J25:J26, K25:K26 (merged pairs = data cells)
    _write(ws, 25, 9, _g(d, 'edad_dias', ''))     # I25 = Dias
    _write(ws, 25, 10, _g(d, 'edad_meses', ''))   # J25 = Meses
    _write(ws, 25, 11, _g(d, 'edad_anios', ''))    # K25 = Anos

    # CUI: O24:S26 merged = data
    cui = _g(d, 'numero_identificacion', _g(d, 'afiliacion', ''))
    tipo_id = _g(d, 'tipo_identificacion', '')
    cui_text = ''
    if tipo_id and cui:
        cui_text = f"{tipo_id}: {cui}"
    elif cui:
        cui_text = cui
    elif tipo_id:
        cui_text = f"{tipo_id}: (sin dato)"
    if cui_text:
        cell = ws.cell(row=24, column=15)
        if not isinstance(cell, MergedCell):
            cell.value = cui_text
            cell.font = Font(name='Calibri', size=8, bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # R27: A27:C29='Nombre del Tutor' (merged label), D27:F29=data(tutor name)
    #      G27:H29='Parentesco' (merged label), I27:K29=data(parentesco)
    #      L27:N28='Codigo Unico de Identificacion' (label), O27:S29=data(CUI tutor)
    # D27:F29 tutor name, I27:K29 parentesco — both can be long
    for col, key in [(4, 'nombre_encargado'), (9, 'parentesco_tutor')]:
        val = _g(d, key)
        if val:
            cell = ws.cell(row=27, column=col)
            if not isinstance(cell, MergedCell):
                cell.value = val
                font_sz = 7 if len(val) > 20 else 8
                cell.font = Font(name='Calibri', size=font_sz, bold=True)
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

    cui_tutor = _g(d, 'numero_id_tutor', '')
    if cui_tutor:
        _write(ws, 27, 15, cui_tutor)  # O27:S29 = CUI tutor data
    # Note: L29:N29 has "(DPI, PASAPORTE, OTRO)" label

    # R30: A30:C30='Pueblo', D30='Maya', E30:F30='Ladino', G30:H30='Garifuna',
    #      I30='Xinca', J30='Extranjero', K30='Si', L30='No',
    #      M30:N30='Migrante', O30='Si', P30='No', Q30='Embarazada', R30='Si', S30='No'
    # Wait — template shows: R30:C4='Maya', C5='Ladino', C7='Garifuna', C9='Xinca',
    #   C10='Extranjero', C11='Si', C12='No', C13='Migrante', C15='Si', C16='No',
    #   C17='Embarazada', C18='Si', C19='No'
    pueblo = _g(d, 'pueblo_etnia', '').upper()
    _check(ws, 30, 4, 'MAYA' in pueblo)                              # D30 = Maya
    _check(ws, 30, 5, 'LADINO' in pueblo or 'MESTIZO' in pueblo)    # E30 = Ladino
    _check(ws, 30, 7, 'GARIF' in pueblo)                             # G30 = Garifuna
    _check(ws, 30, 9, 'XINCA' in pueblo)                             # I30 = Xinca

    pais = _g(d, 'pais_residencia', '').upper()
    es_extranjero = 'EXTRANJERO' in pueblo or (pais and pais not in ('GUATEMALA', 'GT', ''))
    has_pueblo_or_pais = bool(pueblo) or bool(pais)
    _check(ws, 30, 11, es_extranjero)                          # K30 = Si (Extranjero)
    _check(ws, 30, 12, not es_extranjero and has_pueblo_or_pais) # L30 = No (Extranjero)

    migrante = _g(d, 'es_migrante', '')
    _check(ws, 30, 15, _chk(migrante))      # O30 = Si (Migrante)
    _check(ws, 30, 16, _is_no(migrante))     # P30 = No (Migrante)

    embarazada = _g(d, 'esta_embarazada', '')
    _check(ws, 30, 18, _chk(embarazada))     # R30 = Si (Embarazada)
    _check(ws, 30, 19, _is_no(embarazada))   # S30 = No (Embarazada)

    # R31: A31:C31='Trimestre de Embarazo', E31(not merged)=data(trimestre),
    #      G31:I31='Ocupacion'(label+data? or just label?),
    #      L31:N31='Escolaridad', O31:P31=data(escolaridad),
    #      Q31:S31='Telefono' data
    # Actually looking at values: R31:C5='Ocupacion', C10='Escolaridad', C15='Telefono'
    # So col5='Ocupacion' label, col10='Escolaridad' label, col15='Telefono' label
    # Data for trimestre: overwrite the label or use adjacent cell
    trimestre = _g(d, 'trimestre_embarazo')
    if trimestre:
        _write(ws, 31, 4, trimestre)  # D31 = trimestre data (between label and Ocupacion)

    # Ocupacion: G31:I31 is merged = label "Ocupacion". Data in J31 or nearby?
    # Actually looking at the template values: R31:C5='Ocupacion', C10='Escolaridad', C15='Telefono'
    # Those are in columns 5, 10, 15 respectively. These are LABELS.
    # Data goes after each label. For Ocupacion (col 5), data is col 7 (G31:I31 merged = label)
    # Hmm, let me re-examine. R31 merged ranges: A31:C31, G31:I31, L31:N31, O31:P31, Q31:S31
    # Values: C1='Trimestre de Embarazo', C5='Ocupacion', C10='Escolaridad', C15='Telefono'
    # So A31:C31=label(Trimestre), col4=data? col5=label(Ocupacion)
    # G31:I31=label? No — the value 'Ocupacion' is at C5 which is col5 E31.
    # Wait: "R31:C5" means Row31, Column5 = E31. So E31='Ocupacion' (not merged).
    # G31:I31 is merged and EMPTY = data area for Ocupacion.
    # L31:N31 merged and has value? C12='Escolaridad'... no, C12 is col12 = L31.
    # Wait the cell dump says: R31:C10='Escolaridad' = col10 = J31. Not in a merge.
    # But L31:N31 is merged. And O31:P31 is merged.
    # OK let me re-read: R31:C5='Ocupacion' → col5=E31, R31:C10='Escolaridad' → col10=J31
    # R31:C15='Telefono' → col15=O31
    # Merges at R31: A31:C31(label Trimestre), G31:I31(empty=data Ocupacion?),
    #   L31:N31(empty=data Escolaridad?), O31:P31(label Telefono), Q31:S31(data Telefono)
    # G31:I31 Ocupacion, L31:N31 Escolaridad — use wrap for long values
    for col, key in [(7, 'ocupacion'), (12, 'escolaridad')]:
        val = _g(d, key)
        if val:
            cell = ws.cell(row=31, column=col)
            if not isinstance(cell, MergedCell):
                cell.value = val
                font_sz = 7 if len(val) > 15 else 8
                cell.font = Font(name='Calibri', size=font_sz, bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    _write(ws, 31, 17, _g(d, 'telefono_paciente', _g(d, 'telefono_encargado', '')))  # Q31:S31 = Telefono data

    # R32: D32:F32='Pais de Residencia' data?
    # Values: R32:C1='Pais de Residencia', C7='Departamento de Residencia', C13='Municipio de Residencia'
    # = A32=col1, G32=col7, M32=col13
    # Merges: D32:F32, G32:I32, J32:L32, M32:O32, P32:S32
    # So A32 (not merged at R32)='Pais de Residencia' label
    # D32:F32 = data (pais), G32:I32 = label (Departamento), J32:L32 = data (depto)
    # M32:O32 = label (Municipio), P32:S32 = data (municipio)?
    # Wait — R32:C7='Departamento de Residencia' is col7=G32. G32:I32 merged.
    # So G32:I32 = label. Then J32:L32 = data for departamento.
    # M32:O32 merged with value at col13='Municipio de Residencia' — but col13=M32, yes.
    # P32:S32 = data for municipio.
    # D32:F32 Pais, J32:L32 Departamento, P32:S32 Municipio
    for col, key, default in [(4, 'pais_residencia', 'GUATEMALA'), (10, 'departamento_residencia', ''), (16, 'municipio_residencia', '')]:
        val = _g(d, key, default)
        if val:
            cell = ws.cell(row=32, column=col)
            if not isinstance(cell, MergedCell):
                cell.value = val
                cell.font = Font(name='Calibri', size=8, bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # R33: A33:C34='Direccion de Residencia' (merged), D33:L34=data(direccion) (merged)
    #      M33:O34='Lugar Poblado' (merged label), P33:S34=data(poblado) (merged)
    # D33:L34 for address — wide but can have long text. Use wrap.
    addr = _g(d, 'direccion_exacta')
    if addr:
        cell = ws.cell(row=33, column=4)
        if not isinstance(cell, MergedCell):
            cell.value = addr
            cell.font = Font(name='Calibri', size=8, bold=True)
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    # P33:S34 for poblado
    poblado = _g(d, 'poblado')
    if poblado:
        cell = ws.cell(row=33, column=16)
        if not isinstance(cell, MergedCell):
            cell.value = poblado
            cell.font = Font(name='Calibri', size=8, bold=True)
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

    # ===== SECCION 3: ANTECEDENTES MEDICOS Y DE VACUNACION (Rows 35-44) =====

    # R36: A36:C36='Paciente Vacunado', D36='Si', E36='No',
    #      F36:H36='Desconocido/Verbal', J36:L36='Antecedentes medicos',
    #      M36:N36='Si', O36:P36='No', Q36(not merged)='Desconocido'
    # Wait — template values: R36:C4='Si', C5='No', C6='Desconocido/Verbal',
    #   C10='Antecedentes medicos', C13='Si', C15='No', C17='Desconocido'
    vacunado = _g(d, 'vacunado', '')
    _check(ws, 36, 4, _chk(vacunado))                                    # D36 = Si
    _check(ws, 36, 5, _is_no(vacunado))                                  # E36 = No
    _check(ws, 36, 6, _is_desc(vacunado) or 'VERBAL' in str(vacunado).upper())  # F36 = Desc/Verbal

    antec = _g(d, 'tiene_antecedentes_medicos', '')
    _check(ws, 36, 13, _chk(antec))     # M36 = Si
    _check(ws, 36, 15, _is_no(antec))   # O36 = No
    _check(ws, 36, 17, _is_desc(antec)) # Q36 = Desconocido

    # R37: A37:C38='Antecedentes Medicos' (merged), D37:E38='Desnutricion',
    #      G37:I38='Inmunocompromiso', K37:M38='Enfermedad Cronica',
    #      N37:N38(merged)=not clear, Q37:S37='Especifique' data area
    # Values: R37:C4='Desnutricion', C7='Inmunocompromiso', C11='Enfermedad Cronica', C15='Especifique'
    _check(ws, 37, 4, _chk(_g(d, 'antecedente_desnutricion')))       # D37
    _check(ws, 37, 7, _chk(_g(d, 'antecedente_inmunocompromiso')))   # G37
    _check(ws, 37, 11, _chk(_g(d, 'antecedente_enfermedad_cronica'))) # K37

    # Antecedentes checkbox labels are long — shrink to fit in narrow merges
    for r, c in [(37, 4), (37, 7), (37, 11)]:
        cell = ws.cell(row=r, column=c)
        if not isinstance(cell, MergedCell) and cell.value and str(cell.value).startswith('\u2612'):
            cell.font = Font(name='Calibri', size=7, bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center', shrink_to_fit=True)
    antec_detalle = _g(d, 'antecedentes_medicos_detalle', '')
    if antec_detalle:
        cell = ws.cell(row=37, column=17)
        if not isinstance(cell, MergedCell):
            cell.value = antec_detalle
            if len(antec_detalle) <= 22:
                cell.font = Font(name='Calibri', size=7, bold=True)
                cell.alignment = Alignment(horizontal='left', vertical='center')
            else:
                # Long text in constrained height — shrink to fit
                cell.font = Font(name='Calibri', size=7, bold=True)
                cell.alignment = Alignment(horizontal='left', vertical='center', shrink_to_fit=True)
    # O38:S38 merged = additional data area
    antec_extra = _g(d, 'antecedentes_medicos_extra', '')
    if antec_extra:
        _write(ws, 38, 15, antec_extra)

    # R39: A39:D39='Tipo De Vacuna Recibida', E39:G39='Numero De Dosis',
    #      H39:J39='Fecha De La Ultima Dosis', K39:P39='Fuente De La Informacion',
    #      Q39:S39='Vacunacion En El Sector'
    # This is a header row for the vaccine table below.

    # R40-41: SPR vaccine (A40:D41 merged)
    # R40: A40:D41='SPR – Sarampion Paperas Rubeola' (merged label)
    #      E40:G41=data(dosis count), H40='Dia'(label), I40(not in merge)='Mes', J40(not)='Ano'
    #      But wait — looking at merges: E40:G41 merged = dosis data
    #      Q40:R41, S40:S41 = sector checkboxes
    # R40 values: C8='Dia', C9='Mes', C10='Ano', C11='Carne De Vacunacion', C17='MSPAS'
    # So H40='Dia', I40='Mes', J40='Ano' — these are LABELS for fecha ultima dosis
    # Data could go in H41,I41,J41? No — H40 is not in a merge covering 40-41.
    # Actually: no merge covers H40-H41. So H40='Dia' label and H41 would be data...
    # but H41 is not in the cell dump (no value). Let me check merges around row 40-41.
    # Merges: A40:D41, E40:G41, Q40:R41, S40:S41. K40:O40, K41:O41 are separate row merges.
    # So for fecha: H40 is available as a label, and there's no separate data row.
    # The pattern is: overwrite the Dia/Mes/Ano labels with actual values (same as original).
    dosis_spr = _g(d, 'dosis_spr', _g(d, 'numero_dosis_spr', ''))
    _write(ws, 40, 5, dosis_spr)     # E40:G41 = dosis count data

    fecha_spr = _g(d, 'fecha_ultima_spr', _g(d, 'fecha_ultima_dosis', ''))
    _write_date(ws, 40, 8, 9, 10, fecha_spr)  # H40=Dia, I40=Mes, J40=Ano (overwrite labels)

    # Fuente verificacion: K40:O40='Carne De Vacunacion'
    fuente_vac = _g(d, 'fuente_info_vacuna', '').upper()
    sector_vac = _g(d, 'sector_vacunacion', '').upper()
    combined_vac = f"{fuente_vac} {sector_vac}"
    if 'CARNE' in combined_vac or 'CARNÉ' in combined_vac:
        _check(ws, 40, 11, True)  # K40 = Carne checkbox

    # Sector: Q40:R41='MSPAS' checkbox
    if 'MSPAS' in combined_vac:
        _check(ws, 40, 17, True)  # Q40 = MSPAS

    # R41: K41:O41='SIGSA 5a Cuaderno' — checkbox if applicable
    if 'SIGSA' in combined_vac or '5A' in combined_vac:
        _check(ws, 41, 11, True)  # K41

    # R42-43: SR vaccine (A42:D43 merged)
    # E42:G43=dosis data, H42='Dia', I42='Mes', J42='Ano'
    # K42:O42='SIGSA 5B Otros Grupos', Q42:R43='IGSS', S42:S43 merged
    dosis_sr = _g(d, 'dosis_sr', '')
    _write(ws, 42, 5, dosis_sr)    # E42:G43
    _write_date(ws, 42, 8, 9, 10, _g(d, 'fecha_ultima_sr', ''))

    if 'IGSS' in combined_vac:
        _check(ws, 42, 17, True)  # Q42 = IGSS
    if 'SIGSA' in combined_vac or '5B' in combined_vac:
        _check(ws, 42, 11, True)  # K42 = SIGSA 5B

    # R43: K43:O44='Registro Unico de Vacunacion' (merged)
    if 'REGISTRO' in combined_vac or 'UNICO' in combined_vac or 'RUV' in combined_vac:
        _check(ws, 43, 11, True)  # K43

    # R44-45: SPRV vaccine (A44:D45 merged)
    # E44:G45=dosis data, H44(not in dump - no value)
    # Q44:R45='Privado'
    dosis_sprv = _g(d, 'dosis_sprv', '')
    _write(ws, 44, 5, dosis_sprv)  # E44:G45
    _write_date(ws, 44, 8, 9, 10, _g(d, 'fecha_ultima_sprv', ''))

    if 'PRIVADO' in combined_vac:
        _check(ws, 44, 17, True)  # Q44 = Privado

    # K45:O45='another source' if applicable (not in dump, probably empty label)

    # ===== SECCION 4: DATOS CLINICOS (Rows 46-57) =====

    # R47: A47:D48='Fecha de Inicio de Sintomas' (merged), then date labels/data
    # Values: R47:C5='Dia', C6='Mes', C7='Ano' (col5=E47, col6=F47, col7=G47)
    #   R47:C8='Fecha de Inicio de Fiebre'=H47 (H47:J48 merged)
    #   R47:C11='Dia', C12='Mes', C13='Ano'
    #   R47:C14='Fecha de inicio de Exantema/Rash'=N47 (N47:P48 merged)
    #   R47:C17='Dia', C18='Mes', C19='Ano'
    # So dates overwrite Dia/Mes/Ano at: E47,F47,G47 / K47,L47,M47 / Q47,R47,S47
    # Wait — the fiebre Dia is C11=col11=K47. And exantema Dia is C17=col17=Q47.
    _write_date(ws, 47, 5, 6, 7, _g(d, 'fecha_inicio_sintomas'))      # E47,F47,G47
    _write_date(ws, 47, 11, 12, 13, _g(d, 'fecha_inicio_fiebre', '')) # K47,L47,M47
    _write_date(ws, 47, 17, 18, 19, _g(d, 'fecha_inicio_erupcion'))   # Q47,R47,S47

    # R49-52: Symptoms grid (left and right columns)
    # R49: A49='Fiebre'(not in merge? A49 is solo), B49='Si'(solo), C49:D49='Temp. C°',
    #      E49='No'(solo), G49:H49='Desconocido',
    #      J49:L49='Coriza / Catarro', M49='Si'(solo), O49='No'(solo), Q49:R49='Desconocido'
    # Wait — template values: R49:C1='Fiebre'=A49, C2='Si'=B49, C3='Temp. C°'=C49,
    #   C5='No'=E49, C7='Desconocido'=G49,
    #   C10='Coriza / Catarro'=J49, C13='Si'=M49, C15='No'=O49, C17='Desconocido'=Q49
    _check(ws, 49, 2, _chk(_g(d, 'signo_fiebre')))        # B49 = Si
    _check(ws, 49, 5, _is_no(_g(d, 'signo_fiebre')))      # E49 = No
    _check(ws, 49, 7, _is_desc(_g(d, 'signo_fiebre')))    # G49 = Desconocido
    temp = _g(d, 'temperatura_celsius', '')
    if temp:
        _write(ws, 49, 3, f"{temp}\u00b0C")               # C49:D49 = Temp data

    _check(ws, 49, 13, _chk(_g(d, 'signo_coriza')))       # M49 = Si
    _check(ws, 49, 15, _is_no(_g(d, 'signo_coriza')))     # O49 = No
    _check(ws, 49, 17, _is_desc(_g(d, 'signo_coriza')))   # Q49 = Desconocido

    # R50: A50:B50='Exantema/Rash', C50='Si'(solo), E50='No', G50:H50='Desconocido'
    #      J50:L50='Manchas de Koplik', M50='Si', O50='No', Q50:R50='Desconocido'
    # Values: C3='Si'=C50, C5='No'=E50, C7='Desconocido'=G50,
    #   C10='Manchas de Koplik'=J50, C13='Si'=M50, C15='No'=O50, C17='Desconocido'=Q50
    _check(ws, 50, 3, _chk(_g(d, 'signo_exantema')))      # C50 = Si
    _check(ws, 50, 5, _is_no(_g(d, 'signo_exantema')))    # E50 = No
    _check(ws, 50, 7, _is_desc(_g(d, 'signo_exantema')))  # G50 = Desconocido

    _check(ws, 50, 13, _chk(_g(d, 'signo_manchas_koplik')))    # M50 = Si
    _check(ws, 50, 15, _is_no(_g(d, 'signo_manchas_koplik')))  # O50 = No
    _check(ws, 50, 17, _is_desc(_g(d, 'signo_manchas_koplik')))# Q50 = Desconocido

    # R51: A51:B51='Tos', C51='Si'(solo), E51='No', G51:H51='Desconocido'
    #      J51:L51='Artralgia / Artritis', M51='Si', O51='No', Q51:R51='Desconocido'
    _check(ws, 51, 3, _chk(_g(d, 'signo_tos')))           # C51 = Si
    _check(ws, 51, 5, _is_no(_g(d, 'signo_tos')))         # E51 = No
    _check(ws, 51, 7, _is_desc(_g(d, 'signo_tos')))       # G51 = Desconocido

    _check(ws, 51, 13, _chk(_g(d, 'signo_artralgia')))    # M51 = Si
    _check(ws, 51, 15, _is_no(_g(d, 'signo_artralgia')))  # O51 = No
    _check(ws, 51, 17, _is_desc(_g(d, 'signo_artralgia')))# Q51 = Desconocido

    # R52: A52:B52='Conjuntivitis', C52='Si', E52='No', G52:H52='Desconocido'
    #      J52:L52='Adenopatias', M52='Si', O52='No', Q52:R52='Desconocido'
    _check(ws, 52, 3, _chk(_g(d, 'signo_conjuntivitis')))      # C52 = Si
    _check(ws, 52, 5, _is_no(_g(d, 'signo_conjuntivitis')))    # E52 = No
    _check(ws, 52, 7, _is_desc(_g(d, 'signo_conjuntivitis')))  # G52 = Desconocido

    _check(ws, 52, 13, _chk(_g(d, 'signo_adenopatias')))       # M52 = Si
    _check(ws, 52, 15, _is_no(_g(d, 'signo_adenopatias')))     # O52 = No
    _check(ws, 52, 17, _is_desc(_g(d, 'signo_adenopatias')))   # Q52 = Desconocido

    # R53-54: Hospitalizacion (A53:B54 merged)
    # C53:C54='Si'(merged), E53(value? not in dump → probably data or empty)
    # Wait — template values: R53:C1='Hospitalizacion'=A53, C3='Si'=C53, C5='No'=E53,
    #   C7='Desconocido'=G53, C10='Nombre del Hospital'=J53, C14='Fecha de Hospitalizacion'=N53
    #   C17='Dia'=Q53, C18='Mes'=R53, C19='Ano'=S53
    # Merges: A53:B54, C53:C54, D53:D54(not in values=empty), E53 not merged directly...
    # Actually merges show: C53:C54, D53:D54, E53:E54 (not in values → probably empty data cells?)
    # No — E53 has no value in the dump. Let me check: F53:F54, G53:H54, I53:I54, J53:K54, etc.
    # The pattern: C53:C54=Si checkbox area, D53:D54=empty, E53:E54=empty,
    #   F53:F54=empty, G53:H54='Desconocido', I53:I54=empty
    # Hmm, this is complex. Let me just use what we know from the cell dump.
    hosp = _g(d, 'hospitalizado', '')
    _check(ws, 53, 3, _chk(hosp))       # C53 = Si
    _check(ws, 53, 5, _is_no(hosp))     # E53 = No (E53:E54 merged? No, actually F53:F54)
    # Wait — let me use the merges. C53:C54, D53:D54, E53:E54 ... no, the merges show:
    # R53C3:R54C3, R53C4:R54C4, R53C5:R54C5, R53C6:R54C6
    # So C53 through F53 are each merged with row 54. C53='Si' is in the dump.
    # D53, E53, F53 are NOT in the dump = empty cells. D53:D54 might be empty checkbox.
    # The pattern for Hospitalizacion is: C53='Si', then skip... G53:H54='Desconocido'
    # But where is 'No'? Looking at values: C5='No'=E53. Yes! E53='No' is not shown because
    # the dump shows R53:C5='No' — wait, it DOES show it! Let me check again.
    # R53:C5='No' is in the dump. So E53='No'.
    _check(ws, 53, 5, _is_no(hosp))     # E53:E54 = No
    _check(ws, 53, 7, _is_desc(hosp))   # G53:H54 = Desconocido

    # J53:K54='Nombre del Hospital' (merged label)... but R53:C10='Nombre del Hospital'
    # That's col10=J53. And J53:K54 is merged. So where does data go?
    # L53:M54 is merged and empty — this is the data area for hospital name.
    # L53:M54 is very narrow (~10 char widths). Use small font + wrap for hospital name.
    hosp_nombre = _g(d, 'hosp_nombre')
    if hosp_nombre:
        cell = ws.cell(row=53, column=12)
        if not isinstance(cell, MergedCell):
            cell.value = hosp_nombre
            cell.font = Font(name='Calibri', size=6, bold=True)
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

    # Fecha de Hospitalizacion: N53:P54 merged label, Q53=Dia, R53=Mes, S53=Ano
    # But Q53 is not in a special merge (Q53 is solo in row 53).
    # Actually merges show: N53:P54 merged. Q-S are not merged at row 53.
    _write_date(ws, 53, 17, 18, 19, _g(d, 'hosp_fecha', ''))  # Q53=Dia, R53=Mes, S53=Ano

    # R55: A55:B55='Complicaciones', C55(not merged? no, check)
    # Merges at R55: A55:B55, G55:H55, J55:K57(big merge!), L55:M55, Q55:R55
    # Values: R55:C1='Complicaciones'=A55, C3='Si'=C55, C5='No'=E55, C7='Desconocido'=G55,
    #   C10='Especifique Complicaciones'=J55(but J55:K57 is merged = big label area),
    #   C12='Encefalitis'=L55, C15='Diarrea'=O55(solo), C17='Trombocitopenia'=Q55
    comp = _g(d, 'tiene_complicaciones', _g(d, 'complicaciones', ''))
    _check(ws, 55, 3, _chk(comp))       # C55 = Si
    _check(ws, 55, 5, _is_no(comp))     # E55 = No
    _check(ws, 55, 7, _is_desc(comp))   # G55 = Desconocido

    _check(ws, 55, 12, _chk(_g(d, 'comp_encefalitis')))     # L55 = Encefalitis
    _check(ws, 55, 15, _chk(_g(d, 'comp_diarrea')))         # O55 = Diarrea
    _check(ws, 55, 17, _chk(_g(d, 'comp_trombocitopenia'))) # Q55 = Trombocitopenia

    # R56: A56:B56='Aislamiento', C56:C57='Si'(merged), D56:E56='Dia Mes Ano'(date area),
    #      F56:F57='No'(merged), G56:H57='Desconocido'(merged), I56:I57=empty
    #      L56:M57='Otitis Media Aguda'(merged), N56:N57=empty, O56='Neumonia'(solo?),
    #      Q56(not in values)
    # Values: R56:C1='Aislamiento'=A56, C3='Si'=C56, C4='Dia Mes Ano'=D56, C6='No'=F56,
    #   C7='Desconocido'=G56, C12='Otitis Media Aguda'=L56, C15='Neumonia'=O56,
    #   C17='Otras(Especifique)'=Q56
    # R57: A57:B57='Respiratorio', C57 is part of C56:C57 merge,
    #   O57='Ceguera'(? actually R57:C15='Ceguera')
    aisl = _g(d, 'aislamiento_respiratorio', '')
    _check(ws, 56, 3, _chk(aisl))       # C56:C57 = Si
    if _chk(aisl):
        fecha_aisl = _g(d, 'fecha_aislamiento', '')
        dd, mm, yyyy = _parse_date(fecha_aisl)
        if dd:
            _write(ws, 56, 4, f"{dd}/{mm}/{yyyy}")  # D56:E56 = date data
    _check(ws, 56, 6, _is_no(aisl))     # F56:F57 = No
    _check(ws, 56, 7, _is_desc(aisl))   # G56:H57 = Desconocido (but G56 is top-left of merge? checking...)
    # Merge is G56:H57 — so G56 is the top-left. But the value 'Desconocido' is in G56.
    # _check will prepend checkmark to that value. Good.

    # Complicaciones continued
    _check(ws, 56, 12, _chk(_g(d, 'comp_otitis')))    # L56 = Otitis Media Aguda
    _check(ws, 56, 15, _chk(_g(d, 'comp_neumonia')))  # O56 = Neumonia (N56:N57? No, O56 solo)
    # Actually checking: N56:N57 is merged and empty. O56 is not in a merge. But...
    # Let me check: R56:C15='Neumonia'=O56 (not in a merge → it's solo). OK.
    # O56 is written as a solo cell.
    # Q56(not in specific merge at 56, but checking... merge L56:M57 covers L-M.
    # R56:C17='Otras(Especifique)'=Q56. Not in a merge. Solo cell.
    comp_otra = _g(d, 'comp_otra_texto', _g(d, 'complicaciones_otra', ''))
    if comp_otra:
        _check(ws, 56, 17, True)  # Q56 = Otras checkbox
        # Write detail... where? There might be space next to it.
        # Actually the "Otras(Especifique)" label implies data goes after.
        # No specific data cell in R56. Could use R58 area or overwrite.

    # R57: A57:B57='Respiratorio' (continuation of Aislamiento), C57 part of C56:C57 merge
    #   R57:C15='Ceguera'=O57 (solo cell)
    _check(ws, 57, 15, _chk(_g(d, 'comp_ceguera')))   # O57 = Ceguera (R57:C15)

    # Fix narrow complication checkbox cells (O55, O56, O57, Q56) — width ~4.7 chars
    # These solo cells can't fit "☒ Diarrea" etc. at normal font size.
    # Reduce font to 6pt and use shrink_to_fit for these specific cells.
    for r, c in [(55, 15), (56, 15), (57, 15), (56, 17)]:
        cell = ws.cell(row=r, column=c)
        if not isinstance(cell, MergedCell) and cell.value:
            cell.font = Font(name='Calibri', size=6, bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center', shrink_to_fit=True)

    # ===== SECCION 5: FACTORES DE RIESGO (Rows 61-69) =====

    # R62: A62:L62='Existe algun caso confirmado en la comunidad...' (merged)
    #      M62='SI'(solo), O62='NO'(solo), Q62:R62='Desconocido'
    # Values: C13='SI', C15='NO', C17='Desconocido'
    caso_conf = _g(d, 'caso_sospechoso_comunidad_3m', '')
    _check(ws, 62, 13, _chk(caso_conf))     # M62 = SI
    _check(ws, 62, 15, _is_no(caso_conf))   # O62 = NO
    _check(ws, 62, 17, _is_desc(caso_conf)) # Q62 = Desconocido

    # R63: A63:L63='Tuvo contacto con un caso sospechoso...' (merged)
    #      M63='SI', O63='NO', Q63:R63='Desconocido'
    contacto_sosp = _g(d, 'contacto_sospechoso_7_23', '')
    _check(ws, 63, 13, _chk(contacto_sosp))     # M63 = SI
    _check(ws, 63, 15, _is_no(contacto_sosp))   # O63 = NO
    _check(ws, 63, 17, _is_desc(contacto_sosp)) # Q63 = Desconocido

    # R64: A64:G64='Viajo durante los 7-23 dias...' (merged)
    #      H64='Si'(solo), J64:L64='Pais, Departamento y Municipio' label (merged)
    #      M64:Q64=data (merged), R64='NO'(solo? actually S64)
    # Values: C8='Si'=H64, C10='Pais, Departamento y\nMunicipio'=J64, C18='NO'=R64
    viajo = _g(d, 'viajo_7_23_previo', '')
    _check(ws, 64, 8, _chk(viajo))      # H64 = Si
    _check(ws, 64, 18, _is_no(viajo))   # R64 = NO

    viaje_dest = _g(d, 'viaje_pais', _g(d, 'destino_viaje', ''))
    viaje_dep = _g(d, 'viaje_departamento', '')
    viaje_mun = _g(d, 'viaje_municipio', '')
    dest_parts = [p for p in [viaje_dest, viaje_dep, viaje_mun] if p]
    if dest_parts:
        dest_text = ', '.join(dest_parts)
        cell = ws.cell(row=64, column=13)
        if not isinstance(cell, MergedCell):
            cell.value = dest_text
            font_sz = 7 if len(dest_text) > 25 else 8
            cell.font = Font(name='Calibri', size=font_sz, bold=True)
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

    # R65: A65:A66='Fecha de salida'(merged), B65='Dia', C65='Mes', D65='Ano'
    #      E65:E66='Fecha de Entrada'(merged), F65='Dia', G65='Mes', H65='Ano'
    #      I65:J66='Alguna persona de su casa ha viajado?' (merged)
    #      K65='Si'(solo), M65:N66='Fecha de Retorno'(merged), O65='Dia', P65='Mes', Q65='Ano'
    #      R65:R66='NO'(merged? checking: R65:R66 not in merges. Actually S65:S66 is merged)
    # Values: C2='Dia'=B65, C3='Mes'=C65, C4='Ano'=D65,
    #   C5='Fecha de Entrada'=E65, C6='Dia'=F65, C7='Mes'=G65, C8='Ano'=H65,
    #   C9='Alguna persona...'=I65, C11='Si'=K65, C13='Fecha de Retorno'=M65,
    #   C15='Dia'=O65, C16='Mes'=P65, C17='Ano'=Q65, C18='NO'=R65
    _write_date(ws, 65, 2, 3, 4, _g(d, 'viaje_fecha_salida', ''))    # B65,C65,D65
    _write_date(ws, 65, 6, 7, 8, _g(d, 'viaje_fecha_entrada', ''))   # F65,G65,H65

    familiar_viajo = _g(d, 'familiar_viajo_exterior', '')
    _check(ws, 65, 11, _chk(familiar_viajo))   # K65 = Si
    _check(ws, 65, 18, _is_no(familiar_viajo))  # R65 = NO

    _write_date(ws, 65, 15, 16, 17, _g(d, 'familiar_fecha_retorno', ''))  # O65,P65,Q65

    # R67: A67:K67='El Paciente Estuvo En Contacto Con Una Mujer Embarazada?' (merged)
    #      L67='Si'(solo), N67='No'(solo), P67:R67='Desconocido'(merged)
    # Values: C12='Si'=L67, C14='No'=N67, C16='Desconocido'=P67
    contacto_emb = _g(d, 'contacto_embarazada', '')
    _check(ws, 67, 12, _chk(contacto_emb))     # L67 = Si
    _check(ws, 67, 14, _is_no(contacto_emb))   # N67 = No
    _check(ws, 67, 16, _is_desc(contacto_emb)) # P67 = Desconocido

    # R68-69: Fuente posible de contagio
    # R68: A68:C70='Fuente posible de contagio' (merged label)
    #      D68:E68='Contacto en el Hogar', G68:H68='Servicio de Salud',
    #      J68:K68='Institucion Educativa', M68:N68='Espacio Publico', P68:R68='Comunidad'
    # R69: D69:E70='Evento Masivo', G69:H70='Transporte Internacional',
    #      J69:K70='Desconocido', M69:N70='Otro', P69:S69='Otro(Especifique)' data
    fuente_c = _strip_accents(_g(d, 'fuente_posible_contagio', '').upper())
    _check(ws, 68, 4, 'HOGAR' in fuente_c or 'CASA' in fuente_c)        # D68
    _check(ws, 68, 7, 'SALUD' in fuente_c or 'HOSPITAL' in fuente_c)    # G68
    _check(ws, 68, 10, 'EDUCATIVA' in fuente_c or 'ESCUELA' in fuente_c) # J68
    _check(ws, 68, 13, 'PUBLICO' in fuente_c)                            # M68
    _check(ws, 68, 16, 'COMUNIDAD' in fuente_c and 'EDUCATIVA' not in fuente_c) # P68

    _check(ws, 69, 4, 'MASIVO' in fuente_c or 'EVENTO' in fuente_c)     # D69
    _check(ws, 69, 7, 'INTERNACIONAL' in fuente_c or 'TRANSP' in fuente_c) # G69
    _check(ws, 69, 10, 'DESCONOCIDO' in fuente_c or 'DESC' in fuente_c)  # J69
    _check(ws, 69, 13, 'OTRO' in fuente_c and 'DESCONOCIDO' not in fuente_c) # M69
    fuente_c_otro = _g(d, 'fuente_contagio_otro', '')
    if fuente_c_otro:
        _write(ws, 69, 16, fuente_c_otro)  # P69:S69 = Otro(Especifique) data

    # ===== SECCION 6: ACCIONES DE RESPUESTA (Rows 71-76) =====

    # R72: A72:E72='Se realizo busqueda activa institucional (BAI)?' (merged)
    #      F72='Si'(solo), H72='No'(solo? checking...), J72:O72='Numero de casos...'(merged data area)
    #      P72:S72=empty data
    # Values: C6='Si'=F72, C8='No'=H72, C10='Numero de casos sospechosos en BAI'=J72
    bai = _g(d, 'bai_realizada', '')
    _check(ws, 72, 6, _chk(bai))    # F72 = Si
    _check(ws, 72, 8, _is_no(bai))  # H72 = No
    bai_n = _g(d, 'bai_casos_sospechosos', '')
    if bai_n:
        _write(ws, 72, 16, bai_n)   # P72:S72 = data

    # R73: A73:E73='Se realizo BAC?' (merged)
    #      F73='Si', H73='No', J73:O73='Numero de casos...', P73:S73=data
    bac = _g(d, 'bac_realizada', '')
    _check(ws, 73, 6, _chk(bac))    # F73 = Si
    _check(ws, 73, 8, _is_no(bac))  # H73 = No
    bac_n = _g(d, 'bac_casos_sospechosos', '')
    if bac_n:
        _write(ws, 73, 16, bac_n)   # P73:S73 = data

    # R74: A74:E74='Hubo vacunacion de bloqueo en las primeras 48 a 72hrs?' (merged)
    #      F74='Si', H74='No', J74:O74='Se realizo monitoreo rapido de vacunacion?' (merged)
    #      P74='Si'(solo), R74='No'(solo)
    # Values: C6='Si'=F74, C8='No'=H74, C10='Se realizo monitoreo...'=J74,
    #   C16='Si'=P74, C18='No'=R74
    vb = _g(d, 'vacunacion_bloqueo', '')
    _check(ws, 74, 6, _chk(vb))     # F74 = Si
    _check(ws, 74, 8, _is_no(vb))   # H74 = No

    mr = _g(d, 'monitoreo_rapido_vacunacion', '')
    _check(ws, 74, 16, _chk(mr))    # P74 = Si
    _check(ws, 74, 18, _is_no(mr))   # R74 = No

    # R75: A75:E75='Hubo vacunacion con barrido documentado?' (merged)
    #      F75='Si', H75='No', J75:O75=empty data area
    bd = _g(d, 'vacunacion_barrido', '')
    _check(ws, 75, 6, _chk(bd))     # F75 = Si
    _check(ws, 75, 8, _is_no(bd))   # H75 = No

    # R76: A76:C76='Se le administro vitamina A?' (merged)
    #      D76='Si', F76='No', H76='Desconocido'(solo? checking merges...)
    # Values: C4='Si'=D76, C6='No'=F76, C8='Desconocido'=H76,
    #   C10='Numero de dosis de vitamina A recibidas...'=J76
    # J76:N76 merged = label, O76:S76 merged = data
    vit = _g(d, 'vitamina_a_administrada', '')
    _check(ws, 76, 4, _chk(vit))       # D76 = Si
    _check(ws, 76, 6, _is_no(vit))     # F76 = No
    _check(ws, 76, 8, _is_desc(vit))   # H76 = Desconocido
    vit_dosis = _g(d, 'vitamina_a_dosis', '')
    if vit_dosis:
        _write(ws, 76, 15, f"Dosis: {vit_dosis}")  # O76:S76 = data

    # ===== SECCION 7: LABORATORIO (Rows 77-92) =====

    # R78: A78:C78='Tipo de Muestra' (merged), D78(solo)='Suero', F78='Orina',
    #      H78:I78='Hisopado Nasofaringeo'(merged), K78:N78='Por que no se recolecto...'(merged)
    #      O78:S78=data(motivo)
    # Values: C4='Suero'=D78, C6='Orina'=F78, C8='Hisopado Nasofaringeo'=H78,
    #   C11='Por que no se recolecto las 3 muestras?'=K78
    tipo_m = _g(d, 'tipo_muestra', '').upper()
    has_suero = 'SUERO' in tipo_m or bool(_g(d, 'muestra_suero_fecha'))
    has_orina = 'ORINA' in tipo_m or bool(_g(d, 'muestra_orina_fecha'))
    has_hisop = 'HISOP' in tipo_m or bool(_g(d, 'muestra_hisopado_fecha'))
    _check(ws, 78, 4, has_suero)    # D78 = Suero
    _check(ws, 78, 6, has_orina)    # F78 = Orina
    _check(ws, 78, 8, has_hisop)    # H78 = Hisopado

    motivo_no = _g(d, 'motivo_no_3_muestras', _g(d, 'motivo_no_recoleccion', ''))
    if motivo_no:
        _write(ws, 78, 15, motivo_no)  # O78:S78 = data

    # R79-81: Lab table headers
    # R79: A79:A81='No. de Muestra', B79:C81='Tipo de Muestra y Prueba',
    #      D79:F80='Fecha de Toma de Muestra', G79:I80='Fecha de Envio de Muestra',
    #      J79:K81='Virus', L79:N79='Resultado' (header), O79:Q79='Fecha de Resultado',
    #      R79:S84='Secuenciacion' (big merged area!)
    # R80: L80:L81='IgM'(merged), M80:M81='IgG'(merged), N80:N81='Avidez'(merged),
    #      O80:O81='Dia', P80:P81='Mes', Q80:Q81='Ano'
    # R81: D81='Dia', E81='Mes', F81='Ano', G81='Dia', H81='Mes', I81='Ano'
    # (These are column sub-headers for the lab data rows below)

    # Lab data rows (each sample = 2 rows: Sarampion + Rubeola):
    # R82-83: 1a Suero (Anticuerpo)
    #   A82:A83='1a'(merged), B82:C82='Suero'(merged), D82:D83,E82:E83,F82:F83 = fecha toma
    #   G82:G83,H82:H83,I82:I83 = fecha envio, J82:K82='Sarampion', J83:K83='Rubeola'
    #   L82=IgM sar, M82=IgG sar, N82=Avidez sar, O82=Dia res, P82=Mes res, Q82=Ano res
    #   L83=IgM rub, M83=IgG rub, N83=Avidez rub, O83=Dia res rub, P83=Mes res rub, Q83=Ano res rub
    #   B83:C83='(Anticuerpo)'
    # R84-85: 2da Suero — same pattern with A84:A85='2da'
    #   R85:C18='Resultado'=R85(col18)... this is in R79:S84 merge (Secuenciacion area)
    #   Actually R85:C18='Resultado' and R85:C19='Fecha' — these are sub-headers for secuenciacion
    # R86-87: 1a Orina (ARN viral)
    #   L86:N86, L87:N87 = results (merged for wider area since PCR not split IgM/IgG)
    #   O86:O87..Q87 = fecha resultado
    #   R86:R87, S86:S87 = secuenciacion resultado/fecha
    # R88-89: 1a Hisopado Nasofaringeo (ARN viral)
    # R90-91: Otro

    lab_json = _safe_json(_g(d, 'lab_muestras_json', ''))

    # Map slot → (sarampion_row, rubeola_row)
    _SLOT_ROW_MAP = {
        'suero_1': (82, 83),
        'suero_2': (84, 85),
        'orina_1': (86, 87),
        'hisopado_1': (88, 89),
        'otro': (90, 91),
    }

    if lab_json:
        for sample in lab_json:
            if not isinstance(sample, dict):
                continue

            slot = str(sample.get('slot', '')).lower().strip()
            if slot in _SLOT_ROW_MAP:
                row_s, row_r = _SLOT_ROW_MAP[slot]
            else:
                tipo = str(sample.get('tipo_muestra', '')).upper()
                if 'SUERO' in tipo:
                    numero = str(sample.get('numero_muestra', '1'))
                    if numero == '2' or '2' in tipo:
                        row_s, row_r = 84, 85
                    else:
                        row_s, row_r = 82, 83
                elif 'ORINA' in tipo:
                    row_s, row_r = 86, 87
                elif 'HISOP' in tipo:
                    row_s, row_r = 88, 89
                else:
                    row_s, row_r = 90, 91

            # Fecha toma — D col (merged D:F across 2 rows, write at top-left)
            dd, mm, yyyy = _parse_date(sample.get('fecha_toma', ''))
            if dd:
                _write(ws, row_s, 4, f"{dd}/{mm}/{yyyy}")

            # Fecha envio — G col
            dd, mm, yyyy = _parse_date(sample.get('fecha_envio', ''))
            if dd:
                _write(ws, row_s, 7, f"{dd}/{mm}/{yyyy}")

            # Sarampion results (row_s)
            s_igm = sample.get('sarampion_igm', sample.get('resultado_igm', ''))
            s_igg = sample.get('sarampion_igg', sample.get('resultado_igg', ''))
            s_avidez = sample.get('sarampion_avidez', sample.get('resultado_avidez', ''))

            # Rubeola results (row_r)
            r_igm = sample.get('rubeola_igm', '')
            r_igg = sample.get('rubeola_igg', '')
            r_avidez = sample.get('rubeola_avidez', '')

            if row_s <= 85:
                # Suero rows: L/M/N are separate columns (IgM/IgG/Avidez)
                if s_igm:
                    _write(ws, row_s, 12, _lab_val(s_igm))
                if s_igg:
                    _write(ws, row_s, 13, _lab_val(s_igg))
                if s_avidez:
                    _write(ws, row_s, 14, _lab_val(s_avidez))
                if r_igm:
                    _write(ws, row_r, 12, _lab_val(r_igm))
                if r_igg:
                    _write(ws, row_r, 13, _lab_val(r_igg))
                if r_avidez:
                    _write(ws, row_r, 14, _lab_val(r_avidez))
            else:
                # Orina/Hisopado/Otro rows: L:N merged into one cell per row
                # Concatenate all results into L col (top-left of merge)
                s_parts = [_lab_val(v) for v in [s_igm, s_igg, s_avidez] if v]
                if s_parts:
                    _write(ws, row_s, 12, '/'.join(s_parts))
                r_parts = [_lab_val(v) for v in [r_igm, r_igg, r_avidez] if v]
                if r_parts:
                    _write(ws, row_r, 12, '/'.join(r_parts))

            # Fecha resultado — O,P,Q cols on sarampion row
            dd, mm, yyyy = _parse_date(sample.get('fecha_resultado', ''))
            if dd:
                _write(ws, row_s, 15, dd)   # O col
                _write(ws, row_s, 16, mm)   # P col
                _write(ws, row_s, 17, yyyy) # Q col

            # Secuenciacion from sample — for orina/hisopado/otro rows (R86+)
            # These have R:S merged pairs for secuenciacion
            sec = sample.get('secuenciacion', '')
            if sec and row_s >= 86:
                _write(ws, row_s, 18, sec)  # R col

    else:
        # Fallback: use individual fields
        # 1a Suero (R82-83)
        dd, mm, yyyy = _parse_date(_g(d, 'muestra_suero_fecha', ''))
        if dd:
            _write(ws, 82, 4, f"{dd}/{mm}/{yyyy}")
        dd, mm, yyyy = _parse_date(_g(d, 'muestra_suero_fecha_envio', ''))
        if dd:
            _write(ws, 82, 7, f"{dd}/{mm}/{yyyy}")

        res_igm = _g(d, 'resultado_igm_sarampion_suero', _g(d, 'resultado_igm_cualitativo', ''))
        if res_igm:
            _write(ws, 82, 12, _lab_val(res_igm))
        res_igg = _g(d, 'resultado_igg_sarampion_suero', _g(d, 'resultado_igg_cualitativo', ''))
        if res_igg:
            _write(ws, 82, 13, _lab_val(res_igg))

        dd, mm, yyyy = _parse_date(_g(d, 'fecha_resultado_laboratorio', ''))
        if dd:
            _write(ws, 82, 15, dd)
            _write(ws, 82, 16, mm)
            _write(ws, 82, 17, yyyy)

        # 1a Orina (R86-87)
        dd, mm, yyyy = _parse_date(_g(d, 'muestra_orina_fecha', ''))
        if dd:
            _write(ws, 86, 4, f"{dd}/{mm}/{yyyy}")
        pcr_orina = _g(d, 'resultado_pcr_orina', '')
        if pcr_orina:
            _write(ws, 86, 12, _lab_val(pcr_orina))

        # 1a Hisopado (R88-89)
        dd, mm, yyyy = _parse_date(_g(d, 'muestra_hisopado_fecha', ''))
        if dd:
            _write(ws, 88, 4, f"{dd}/{mm}/{yyyy}")
        pcr_hisop = _g(d, 'resultado_pcr_hisopado', '')
        if pcr_hisop:
            _write(ws, 88, 12, _lab_val(pcr_hisop))

    # Secuenciacion global result and fecha
    # R85:C18='Resultado', R85:C19='Fecha' — but these are inside R79:S84 merge.
    # For suero samples, secuenciacion data goes in R82-85 area.
    # For standalone fields, write them at the first suero secuenciacion spot.
    sec_resultado = _g(d, 'secuenciacion_resultado', '')
    sec_fecha = _g(d, 'secuenciacion_fecha', '')
    # R86-91 have dedicated R:S columns for secuenciacion.
    # For suero (R82-85), the R79:S84 merge covers them — data goes in R79 col18 (top-left of merge).
    # But R79:S84 is labeled 'Secuenciacion'. We'd need to overwrite.
    # Better: don't fight the big merge. Write secuenciacion in the Otro row if not already used.
    if sec_resultado and not lab_json:
        # Write in R90 (Otro) secuenciacion area — R90:R91 merged
        _write(ws, 90, 18, sec_resultado)
    if sec_fecha and not lab_json:
        dd, mm, yyyy = _parse_date(sec_fecha)
        if dd:
            _write(ws, 90, 19, f"{dd}/{mm}/{yyyy}")

    # ===== CLASIFICACION (Rows 93-108) =====

    # R94: A94:C94='Clasificacion Final' (merged), D94:E94='Sarampion'(merged),
    #      G94:H94='Rubeola'(merged), J94:K94='Descartado'(merged),
    #      M94:N94='Pendiente'(merged), P94:R94='No cumple def. caso'(merged)
    clasif = _strip_accents(_g(d, 'clasificacion_caso', '').upper())
    is_confirmed = 'CONFIRM' in clasif
    is_descartado = 'DESCART' in clasif
    is_pendiente = 'PENDIENTE' in clasif or 'SOSPECH' in clasif
    is_no_cumple = 'NO CUMPLE' in clasif or 'DEFINICION' in clasif
    is_clasif_sar = ('SARAMP' in clasif and not is_descartado) or (is_confirmed and not is_descartado and is_sar)
    is_clasif_rub = ('RUBEO' in clasif or 'RUBE' in clasif) and not is_descartado
    if is_confirmed and not is_clasif_sar and not is_clasif_rub and not is_descartado:
        is_clasif_sar = is_sar
        is_clasif_rub = is_rub
    _check(ws, 94, 4, is_clasif_sar)                      # D94 = Sarampion
    _check(ws, 94, 7, is_clasif_rub)                       # G94 = Rubeola
    _check(ws, 94, 10, is_descartado)                      # J94 = Descartado
    _check(ws, 94, 13, is_pendiente and not is_confirmed)  # M94 = Pendiente
    _check(ws, 94, 16, is_no_cumple)                       # P94 = No cumple

    # Clasificacion cells are narrow merges — shrink checked labels
    for r, c in [(94, 4), (94, 7), (94, 10), (94, 13), (94, 16)]:
        cell = ws.cell(row=r, column=c)
        if not isinstance(cell, MergedCell) and cell.value and str(cell.value).startswith('\u2612'):
            cell.font = Font(name='Calibri', size=7, bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center', shrink_to_fit=True)

    # R95: A95:C95='Criterio de Confirmacion', D95:E95='Laboratorio',
    #      G95:H95='Nexo epidemiologico', J95:K95='Clinico',
    #      M95:O95='Contacto de Otro Caso', P95='Si'(solo?), R95(not in dump)
    # Values: C4='Laboratorio'=D95, C7='Nexo epidemiologico'=G95, C10='Clinico'=J95,
    #   C13='Contacto de Otro Caso'=M95, C16='Si'=P95, C18='NO'=R95
    crit_conf = _strip_accents(_g(d, 'criterio_confirmacion', '').upper())
    _check(ws, 95, 4, 'LABORATORIO' in crit_conf or 'LAB' in crit_conf)        # D95
    _check(ws, 95, 7, 'NEXO' in crit_conf or 'EPIDEMIOL' in crit_conf)         # G95
    _check(ws, 95, 10, 'CLINICO' in crit_conf)                                  # J95

    # Shrink checked criterio labels to fit narrow cells
    for r, c in [(95, 4), (95, 7), (95, 10)]:
        cell = ws.cell(row=r, column=c)
        if not isinstance(cell, MergedCell) and cell.value and str(cell.value).startswith('\u2612'):
            cell.font = Font(name='Calibri', size=7, bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center', shrink_to_fit=True)

    contacto_caso = _g(d, 'contacto_otro_caso', '')
    _check(ws, 95, 16, _chk(contacto_caso))     # P95 = Si
    _check(ws, 95, 18, _is_no(contacto_caso))   # R95 = NO

    # R96-97: Criterio para descartar (A96:C97 merged)
    # D96:E97='Laboratorio', G96:H97='Relacionado con la Vacuna',
    # J96:K97='Clinico', M96:O96='Otro Especificar:', P96:S96=data
    # Values: C4='Laboratorio'=D96, C7='Relacionado con la Vacuna'=G96,
    #   C10='Clinico'=J96, C13='Otro Especificar:'=M96
    crit_desc = _strip_accents(_g(d, 'criterio_descarte', '').upper())
    _check(ws, 96, 4, 'LABORAT' in crit_desc)                                  # D96
    _check(ws, 96, 7, 'VACUNA' in crit_desc)                                   # G96
    _check(ws, 96, 10, 'CLINICO' in crit_desc)                                  # J96
    crit_desc_otro = _g(d, 'criterio_descarte_otro', '')
    if crit_desc_otro:
        _write(ws, 96, 16, crit_desc_otro)  # P96:S96 = data

    # R98-99: Fuentes de infeccion de los casos confirmados
    # A98:C99='Fuentes de infeccion...' (merged)
    # D98:E99='Importado', G98:I99='Relacionado con la importacion',
    # J98:L98='Pais de Importacion'(label), M98:M99(merged)=data? or N98:O99=data?
    # P98:P99='Endemico', Q98:R99(merged)=?, S98:S99='Fuente desconocida'
    # Values: C4='Importado'=D98, C7='Relacionado con la importacion'=G98,
    #   C10='Pais de Importacion'=J98, C14='Endemico'=N98, C17='Fuente desconocida'=Q98
    # Wait — the dump says: R98:C14='Endemico' → col14=N98. And R98:C17='Fuente desconocida'=Q98.
    # Merges: N98:O99, P98:P99, Q98:R99, S98:S99
    # So N98:O99='Endemico' checkbox area? No — looking at values, N98 is not in dump.
    # R98:C14='Endemico' = col14 = N98. And N98:O99 is merged. So that's the Endemico checkbox.
    # Q98:R99='Fuente desconocida' area. And S98:S99=empty.
    # Pais de importacion data: J98:L98 is the label merge.
    # M98:M99 is merged and empty — but wait, R99:C10:R99C12 (J99:L99) is merged too.
    # The data for pais de importacion should go in M98 (M98:M99 merged).
    fuente_inf = _strip_accents(_g(d, 'fuente_infeccion', '').upper())
    _check(ws, 98, 4, 'IMPORTADO' in fuente_inf and 'RELACIONADO' not in fuente_inf) # D98
    _check(ws, 98, 7, 'RELACIONADO' in fuente_inf)                                    # G98
    _check(ws, 98, 14, 'ENDEMI' in fuente_inf)                                        # N98
    _check(ws, 98, 17, 'DESCONOCID' in fuente_inf)                                    # Q98

    pais_import = _g(d, 'pais_importacion', '')
    if pais_import:
        _write(ws, 98, 13, pais_import)  # M98:M99 = Pais de importacion data

    # R100: A100:C100='Caso Analizado Por', D100:E100='CONAPI', G100:H100='DEGR*',
    #       J100:K100='Comision Nacional**', M100='Otros'(solo?), O100:R100='Especifique'(merged data)
    # Values: C4='CONAPI'=D100, C7='DEGR*'=G100, C10='Comision Nacional**'=J100,
    #   C13='Otros'=M100, C15='Especifique'=O100
    analizado = _strip_accents(_g(d, 'caso_analizado_por', '').upper())
    _check(ws, 100, 4, 'CONAPI' in analizado)    # D100
    _check(ws, 100, 7, 'DEGR' in analizado)      # G100
    _check(ws, 100, 10, 'COMISION' in analizado or 'NACIONAL' in analizado) # J100
    _check(ws, 100, 13, 'OTRO' in analizado)     # M100
    otro_analiz = _g(d, 'caso_analizado_por_otro', '')
    if otro_analiz:
        _write(ws, 100, 15, otro_analiz)  # O100:R100 = data

    # R101-102: Fecha de Clasificacion + Condicion final del Paciente
    # A101:C102='Fecha de Clasificacion' (merged), D101='Dia'(solo? checking merges...)
    # Actually looking at values: R101:C4='Dia'=D101, C5='Mes'=E101, C6='Ano'=F101
    # Merges: A101:C102. G101:H102='Condicion final del Paciente'(merged label)
    # I101:I102='Recuperado' (merged), K101:L102='Con Secuelas' (merged),
    # N101:O102='Fallecido*' (merged), Q101:R102='Desconocido' (merged)
    # Wait values: C7='Condicion final del Paciente'=G101, C9='Recuperado'=I101,
    #   C11='Con Secuelas'=K101, C14='Fallecido*'=N101, C17='Desconocido'=Q101
    dd, mm, yyyy = _parse_date(_g(d, 'fecha_clasificacion_final', ''))
    if dd:
        _write(ws, 101, 4, dd)   # D101
        _write(ws, 101, 5, mm)   # E101
        _write(ws, 101, 6, yyyy) # F101

    cond = _strip_accents(_g(d, 'condicion_final_paciente', _g(d, 'condicion_egreso', '')).upper())
    _check(ws, 101, 9, 'RECUPER' in cond or 'VIVO' in cond or 'MEJORAD' in cond or 'CURAD' in cond) # I101
    _check(ws, 101, 11, 'SECUELA' in cond)                                                           # K101
    _check(ws, 101, 14, 'FALLEC' in cond or 'MUERTE' in cond)                                        # N101
    _check(ws, 101, 17, 'DESCONOCID' in cond)                                                        # Q101

    # Condicion final cells are narrow merged areas — shrink checked labels to fit
    for r, c in [(101, 9), (101, 11), (101, 14), (101, 17)]:
        cell = ws.cell(row=r, column=c)
        if not isinstance(cell, MergedCell) and cell.value and str(cell.value).startswith('\u2612'):
            cell.font = Font(name='Calibri', size=7, bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center', shrink_to_fit=True)

    # R103-104: Fecha de Defuncion + Causa de Muerte
    # A103:C104='Fecha de Defuncion*' (merged), D103='Dia', E103='Mes', F103='Ano'
    # G103:I104='Causa de Muerte Segun Certificado de Defuncion' (merged label)
    # J103:S103=data (merged), J104:S104=data line 2 (merged)
    dd, mm, yyyy = _parse_date(_g(d, 'fecha_defuncion', ''))
    if dd:
        _write(ws, 103, 4, dd)   # D103
        _write(ws, 103, 5, mm)   # E103
        _write(ws, 103, 6, yyyy) # F103
    _write(ws, 103, 10, _g(d, 'causa_muerte_certificado', ''))  # J103:S103 = data

    # R105-106: Observaciones
    # A105:C106='Observaciones' (merged label), D105:S105=data line 1, D106:S106=data line 2
    obs = _g(d, 'observaciones', '')
    if obs:
        # D105:S105 is wide enough for most observations with wrap_text
        cell = ws.cell(row=105, column=4)
        if not isinstance(cell, MergedCell):
            cell.value = obs
            cell.font = Font(name='Calibri', size=8, bold=True)
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)


# ---------------------------------------------------------------------------
# Excel to PDF conversion via LibreOffice
# ---------------------------------------------------------------------------

def _xlsx_to_pdf(xlsx_path: str) -> bytes:
    """Convert an .xlsx file to PDF using LibreOffice in headless mode."""
    out_dir = os.path.dirname(xlsx_path)

    cmd = [
        "libreoffice",
        "--headless",
        "--calc",
        "--convert-to", "pdf",
        "--outdir", out_dir,
        xlsx_path,
    ]

    try:
        result = subprocess.run(
            cmd,
            capture_output=True,
            text=True,
            timeout=60,
            env={**os.environ, "HOME": "/tmp"},
        )
        if result.returncode != 0:
            logger.error("LibreOffice conversion failed: %s", result.stderr)
            raise RuntimeError(f"LibreOffice conversion failed: {result.stderr}")
    except FileNotFoundError:
        # LibreOffice not installed — try soffice
        cmd[0] = "soffice"
        result = subprocess.run(
            cmd,
            capture_output=True,
            text=True,
            timeout=60,
            env={**os.environ, "HOME": "/tmp"},
        )
        if result.returncode != 0:
            raise RuntimeError(f"soffice conversion failed: {result.stderr}")

    pdf_path = xlsx_path.rsplit(".", 1)[0] + ".pdf"
    if not os.path.exists(pdf_path):
        raise RuntimeError(f"PDF not generated at {pdf_path}. stdout={result.stdout}")

    with open(pdf_path, "rb") as f:
        return f.read()


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def generar_ficha_v2_pdf(data: dict) -> bytes:
    """Generate a single PDF ficha from patient data using the Excel template.

    Args:
        data: dict with patient fields (from database record)

    Returns:
        PDF file contents as bytes
    """
    if not os.path.exists(TEMPLATE_PATH):
        raise FileNotFoundError(f"Template not found: {TEMPLATE_PATH}")

    with tempfile.TemporaryDirectory(prefix="ficha_v2_") as tmpdir:
        # Copy template
        registro_id = _g(data, 'registro_id', 'unknown')
        safe_id = registro_id.replace('/', '_').replace('\\', '_')[:50]
        xlsx_path = os.path.join(tmpdir, f"ficha_{safe_id}.xlsx")
        shutil.copy2(TEMPLATE_PATH, xlsx_path)

        # Open and fill
        wb = openpyxl.load_workbook(xlsx_path)
        ws = wb["Hoja1"]

        _fill_template(ws, data)

        # Page break after Section 4 (row 57 = end of clinical data, before Section 5)
        rb = RowBreak()
        rb.append(Break(id=57))
        ws.row_breaks = rb

        # Print settings: fit width to 1 page, let height flow naturally (2+ pages)
        ws.page_setup.orientation = "portrait"
        ws.page_setup.paperSize = 1  # Letter
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.page_setup.scale = None  # Let fitToPage control sizing
        ws.print_area = "A1:S108"

        wb.save(xlsx_path)

        # Convert to PDF
        pdf_bytes = _xlsx_to_pdf(xlsx_path)

    return pdf_bytes


def generar_fichas_v2_bulk(records: list, merge: bool = True) -> bytes:
    """Generate PDFs for multiple records.

    Args:
        records: list of dicts with patient data
        merge: if True, return a single merged PDF; if False, return a ZIP of individual PDFs

    Returns:
        PDF bytes (if merge=True) or ZIP bytes (if merge=False)
    """
    if not records:
        raise ValueError("No records provided")

    if len(records) == 1:
        return generar_ficha_v2_pdf(records[0])

    with tempfile.TemporaryDirectory(prefix="fichas_v2_bulk_") as tmpdir:
        pdf_files = []

        for i, rec in enumerate(records):
            registro_id = _g(rec, 'registro_id', f'rec_{i}')
            safe_id = registro_id.replace('/', '_').replace('\\', '_')[:50]
            xlsx_path = os.path.join(tmpdir, f"ficha_{i:04d}_{safe_id}.xlsx")
            shutil.copy2(TEMPLATE_PATH, xlsx_path)

            wb = openpyxl.load_workbook(xlsx_path)
            ws = wb["Hoja1"]

            _fill_template(ws, rec)

            # Page break after Section 4
            rb = RowBreak()
            rb.append(Break(id=57))
            ws.row_breaks = rb

            # Print settings
            ws.page_setup.orientation = "portrait"
            ws.page_setup.paperSize = 1
            ws.page_setup.fitToWidth = 1
            ws.page_setup.fitToHeight = 0
            ws.sheet_properties.pageSetUpPr.fitToPage = True
            ws.page_setup.scale = None
            ws.print_area = "A1:S108"

            wb.save(xlsx_path)

            pdf_bytes = _xlsx_to_pdf(xlsx_path)
            pdf_path = os.path.join(tmpdir, f"ficha_{i:04d}_{safe_id}.pdf")
            with open(pdf_path, "wb") as f:
                f.write(pdf_bytes)
            pdf_files.append(pdf_path)

        if merge:
            try:
                from pypdf import PdfReader, PdfWriter
                writer = PdfWriter()
                for pf in pdf_files:
                    reader = PdfReader(pf)
                    for page in reader.pages:
                        writer.add_page(page)
                buf = io.BytesIO()
                writer.write(buf)
                return buf.getvalue()
            except ImportError:
                pass

            try:
                from PyPDF2 import PdfReader, PdfWriter
                writer = PdfWriter()
                for pf in pdf_files:
                    reader = PdfReader(pf)
                    for page in reader.pages:
                        writer.add_page(page)
                buf = io.BytesIO()
                writer.write(buf)
                return buf.getvalue()
            except ImportError:
                pass

            # Last resort: return first PDF only (lossy)
            logger.warning("No PDF merger available; returning only first PDF")
            with open(pdf_files[0], "rb") as f:
                return f.read()
        else:
            # Return ZIP
            buf = io.BytesIO()
            with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
                for pf in pdf_files:
                    zf.write(pf, os.path.basename(pf))
            return buf.getvalue()


# ---------------------------------------------------------------------------
# CLI test
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    test_data = {
        'registro_id': 'TEST-2026-001',
        'diagnostico_sospecha': 'SARAMPION',
        'diagnostico_registrado': 'SARAMPION',
        'fecha_notificacion': '2026-03-15',
        'area_salud_mspas': 'GUATEMALA CENTRAL',
        'distrito_salud_mspas': 'GUATEMALA',
        'servicio_salud_mspas': 'IGSS ZONA 9',
        'fecha_captacion': '2026-03-14',
        'fecha_visita_domiciliaria': '2026-03-16',
        'nom_responsable': 'Dr. Juan Perez',
        'cargo_responsable': 'Medico Epidemiologo',
        'telefono_responsable': '5555-1234',
        'correo_responsable': 'jperez@igss.gob.gt',
        'es_seguro_social': 'SI',
        'unidad_medica': 'Hospital General de Enfermedades IGSS',
        'fuente_notificacion': 'Servicio De Salud',
        'nombres': 'Maria Isabel',
        'apellidos': 'Garcia Lopez',
        'sexo': 'F',
        'fecha_nacimiento': '1995-06-15',
        'edad_anios': '30',
        'edad_meses': '9',
        'numero_identificacion': '1234567890101',
        'afiliacion': '12345678',
        'pueblo_etnia': 'Ladino',
        'pais_residencia': 'Guatemala',
        'departamento_residencia': 'Guatemala',
        'municipio_residencia': 'Guatemala',
        'direccion_exacta': '5a. Avenida 12-34, Zona 1',
        'poblado': 'Ciudad de Guatemala',
        'ocupacion': 'Enfermera',
        'escolaridad': 'Universitaria',
        'telefono_paciente': '5555-5678',
        'esta_embarazada': 'NO',
        'vacunado': 'SI',
        'dosis_spr': '2',
        'fecha_ultima_spr': '2010-03-15',
        'fuente_info_vacuna': 'MSPAS',
        'tiene_antecedentes_medicos': 'NO',
        'fecha_inicio_sintomas': '2026-03-10',
        'fecha_inicio_fiebre': '2026-03-10',
        'fecha_inicio_erupcion': '2026-03-12',
        'signo_fiebre': 'SI',
        'temperatura_celsius': '38.5',
        'signo_exantema': 'SI',
        'signo_tos': 'SI',
        'signo_conjuntivitis': 'SI',
        'signo_coriza': 'SI',
        'signo_manchas_koplik': 'NO',
        'signo_artralgia': 'NO',
        'signo_adenopatias': 'SI',
        'hospitalizado': 'SI',
        'hosp_nombre': 'Hospital General de Enfermedades IGSS',
        'hosp_fecha': '2026-03-14',
        'tiene_complicaciones': 'NO',
        'aislamiento_respiratorio': 'SI',
        'fecha_aislamiento': '2026-03-14',
        'caso_sospechoso_comunidad_3m': 'NO',
        'contacto_sospechoso_7_23': 'SI',
        'viajo_7_23_previo': 'NO',
        'contacto_embarazada': 'NO',
        'fuente_posible_contagio': 'Serv. Salud',
        'bai_realizada': 'SI',
        'bai_casos_sospechosos': '3',
        'bac_realizada': 'NO',
        'vacunacion_bloqueo': 'SI',
        'monitoreo_rapido_vacunacion': 'SI',
        'vacunacion_barrido': 'NO',
        'vitamina_a_administrada': 'SI',
        'vitamina_a_dosis': '1',
        'muestra_suero': 'SI',
        'muestra_suero_fecha': '2026-03-15',
        'muestra_hisopado': 'SI',
        'muestra_hisopado_fecha': '2026-03-15',
        'muestra_orina': 'SI',
        'muestra_orina_fecha': '2026-03-16',
        'resultado_igm_cualitativo': 'POSITIVO',
        'resultado_igg_cualitativo': 'NEGATIVO',
        'resultado_pcr_hisopado': 'POSITIVO',
        'resultado_pcr_orina': 'NEGATIVO',
        'fecha_resultado_laboratorio': '2026-03-18',
        'clasificacion_caso': 'SARAMPION CONFIRMADO',
        'criterio_confirmacion': 'Laboratorio',
        'contacto_otro_caso': 'SI',
        'fuente_infeccion': 'Importado',
        'pais_importacion': 'Honduras',
        'caso_analizado_por': 'CONAPI',
        'fecha_clasificacion_final': '2026-03-20',
        'condicion_final_paciente': 'Recuperado',
        'observaciones': 'Caso importado de Honduras. Contacto con caso confirmado en San Pedro Sula.',
    }

    print("Generating test PDF with full data...")
    pdf_bytes = generar_ficha_v2_pdf(test_data)
    out_path = "/tmp/ficha_v2_test_full.pdf"
    with open(out_path, "wb") as f:
        f.write(pdf_bytes)
    print(f"  Written to {out_path} ({len(pdf_bytes)} bytes)")

    print("Generating test PDF with empty data...")
    pdf_bytes_empty = generar_ficha_v2_pdf({'registro_id': 'EMPTY-TEST'})
    out_path_empty = "/tmp/ficha_v2_test_empty.pdf"
    with open(out_path_empty, "wb") as f:
        f.write(pdf_bytes_empty)
    print(f"  Written to {out_path_empty} ({len(pdf_bytes_empty)} bytes)")

    print("Done.")
